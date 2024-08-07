from django.shortcuts import render
from reactivos.views import *
from .models import *
from reactivos.models import SolicitudesExternas
from reactivos.models import *
from django.views.generic import ListView, View, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.contrib import messages
from django.http import HttpResponseRedirect,JsonResponse, HttpResponseBadRequest, HttpResponse
from django.urls import reverse
from datetime import datetime, timedelta, date
from django.utils.timezone import make_aware
from django.db.models import Q
from django.utils.http import urlencode
from urllib.parse import urlencode
from django.shortcuts import get_object_or_404, redirect
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from django.contrib.staticfiles import finders
from PIL import Image as PILImage
from django.utils.timezone import make_aware
from django.db.models import Q
from datetime import datetime
import os
from .forms import *
import base64
import pytz
from django.db.models import F, ExpressionWrapper, DecimalField
from django.template.loader import get_template
from .utils import *
from django.db.models import Max, Min
import glob
from openpyxl.worksheet.page import PageMargins


# ---------------------------------------------------------- 
# Función para validar usuarios con acceso
def check_group_permission(groups_required):
    def decorator(view_func):
        @method_decorator(login_required)
        def _wrapped_view(self, request, *args, **kwargs):
            grupos_usuario = self.request.user.groups.all().values('name')
            is_of_group = False
            
            for grupo in grupos_usuario:
                for grupo_requerido in groups_required:
                    if grupo['name'] == grupo_requerido:
                        is_of_group = True
                        break
            
            if is_of_group or self.request.user.is_superuser:
                return view_func(self, request, *args, **kwargs)
            else:
                messages.error(request, 'No tiene permisos para acceder a esta vista, desea acceder con credenciales distintas')
                return HttpResponseRedirect(reverse('reactivos:login'))

        return _wrapped_view
    return decorator



# ---------------------------------------------------------- #
# Vista para la creación página principal de aplicativo de residuos,
class HomeUniCLabResiduos(View):  # Utiliza LoginRequiredMixin como clase base
    template_name = 'UniCLab_Residuos/index.html'  # Nombre de la plantilla

    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request,*args,**kwargs):
        
             
        context = {
            
        }
        return render(request, self.template_name, context)


# ---------------------------------------------------------- #
# Listado de Clasificación de residuos
class Wastes_Classification_List(LoginRequiredMixin,ListView):
    model = CLASIFICACION_RESIDUOS
    template_name = "UniCLab_Residuos/listado_clasificaciones.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        
        # Obtener las fechas de inicio y fin de la solicitud GET
        name = self.request.GET.get('name')
        id_classification = self.request.GET.get('id_classification')
           
        # Guardar los valores de filtrado en la sesión
        
        request.session['filtered_name'] = name
        request.session['filtered_id_classification'] = id_classification
        
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['usuarios'] = User.objects.all()
        context['laboratorios'] = Laboratorios.objects.all()
        
       
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()

        id_classification = self.request.GET.get('id_classification')
        if self.request.user.rol.name=='ADMINISTRADOR' or self.request.user.rol.name=='ADMINISTRADOR AMBIENTAL':
            
            if id_classification:
                queryset = queryset.filter(id=id_classification) 
        else:
            
            queryset = queryset.filter(is_active=True)
            if id_classification:
                queryset = queryset.filter(id=id_classification, is_active=True)

        queryset = queryset.order_by('id')
        return queryset
    
# ------------------------------------------------ #
# Paginación de listado de clasifición de residuos #
class Wastes_Classification_Pagination(LoginRequiredMixin,View):
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        
        filtered_name = request.session.get('filtered_name')
        filtered_id_classification = request.session.get('filtered_id_classification')
        
        url = reverse('residuos:clasificacion_residuos')
        params = {}
        
        if filtered_name:
            params['name'] = filtered_name
        if filtered_id_classification:
            params['id_classification'] = filtered_id_classification
                
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)


# --------------------------------- #
# Autocompletado de clasificaciones #

class AutocompleteClassification(LoginRequiredMixin,View):
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request):
        term = request.GET.get('term', '')
        
        clasificaciones = CLASIFICACION_RESIDUOS.objects.filter(
                Q(name__icontains=term) | Q(description__icontains=term)).order_by('description')[:20]
        

        results = []
        for clasificacion in clasificaciones:
            result = {
                'name': clasificacion.name,
                'description': clasificacion.description,
                'id_classification':clasificacion.id,
            }
            results.append(result)
        return JsonResponse(results, safe=False)
    
# ------------------------------------------------------- #
# Exportar a Excel listado de Clasiifcaciones de residuos #

class ExportToExcelClassification(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    

    
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        def get_file_name():
            now = datetime.now()
            return f'Clasificacion_residuos_decreto_1076_2015_{now.strftime("%d%m%Y%H%M%S")}.xlsx'
        id_classification = request.session.get('filtered_id_classification')
        queryset=CLASIFICACION_RESIDUOS.objects.all()
        if id_classification:
            queryset = queryset.filter(id=id_classification)

        queryset = queryset.order_by('id')

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')
        pil_image = PILImage.open(logo_path)
        image = ExcelImage(pil_image)
        sheet.add_image(image, 'A1')

        fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        sheet['B1'] = 'Clasificación de residuos según decreto 2076 de 2015'
        sheet['B2'] = 'Fecha de Creación: '+fecha_creacion
        sheet['A4'] = 'Nombre'
        sheet['B4'] = 'Descripción'
        sheet['C4'] = 'Estado'

        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 30
        sheet.row_dimensions[3].height = 25

        Titulo = sheet['B1']
        Titulo.font = Font(bold=True, size=16)
        Titulo.alignment = Alignment(horizontal='right')

        Fecha = sheet['B2']
        Fecha.alignment = Alignment(horizontal='right')
        
        

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        bold_font = Font(bold=True)

        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 125

        alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        

        row = 4
        for col in range(1, 4):
            sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).font = bold_font

        row = 5
        for item in queryset:
            if item.is_active:
                Estado="Activo"
            else:
                Estado="Inactivo"
            sheet.cell(row=row, column=1).value = (item.name)
            sheet.cell(row=row, column=2).value = (item.description)
            sheet.cell(row=row, column=3).value = Estado
            

            for col in range(1, 4):
                sheet.cell(row=row, column=col).border = thin_border
                sheet.cell(row=row, column=col).alignment = alignment

            row += 1

        start_column = 1
        end_column = 3
        start_row = 4
        end_row = row - 1

        start_column_letter = get_column_letter(start_column)
        end_column_letter = get_column_letter(end_column)

        table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

        sheet.auto_filter.ref = table_range

        fill = PatternFill(fill_type="solid", fgColor=WHITE)
        start_cell = sheet['A1']
        end_column_letter = get_column_letter(end_column+1)
        end_row = row+1
        end_cell = sheet[end_column_letter + str(end_row)]
        table_range = start_cell.coordinate + ':' + end_cell.coordinate

        for row in sheet[table_range]:
            for cell in row:
                cell.fill = fill

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={get_file_name()}'

        workbook.save(response)

        tipo_evento = 'DESCARGA DE ARCHIVOS'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        return response
    
# ------------------------------- #
# Crear Clasificación de Residuos #

class CrateWasteClassification(LoginRequiredMixin ,View):
    template_name = 'UniCLab_Residuos/crear_clasificacion.html'

    
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        form = ClasificacionResiduosForm()
        return render(request, self.template_name, {'form': form})
    
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        form = ClasificacionResiduosForm(request.POST, request.FILES)
        
        try:
            if form.is_valid():
                # Asignar el usuario actual a los campos created_by y last_updated_by
                form.instance.created_by = request.user
                form.instance.last_updated_by = request.user
                # Guardar la clasificación si el formulario es válido
                clasificacion = form.save()
                
                # Registrar evento
                tipo_evento = 'CREAR CLASIFICACION DE RESIDUOS'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)
                
                mensaje=f'Clasificación de residuos creada correctamente.'
                # Devuelve una respuesta JSON de éxito
                return JsonResponse({'success': True, 'message': mensaje})
                
                
            else:
                # Devuelve una respuesta JSON con los errores de validación
                return JsonResponse({'success': False, 'errors': form.errors})
            
        except Exception as e:
            # Imprimir el error en la consola o logs
            print(e)
            mensaje=f'Error: {e}'
            # Devolver una respuesta de error o redirigir a una página de error
            return HttpResponseBadRequest(f'Error interno del servidor:{mensaje}')
   
# -------------------------------- #
# Editar Clasificación de Residuos #

class EditWasteClassification(LoginRequiredMixin, View):
    template_name = 'UniCLab_Residuos/editar_clasificacion.html'

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        try:
            # Decodificar el ID en base64 dos veces
            class_key = kwargs.get('pk')
            classification_id = base64.urlsafe_b64decode(base64.urlsafe_b64decode(class_key)).decode('utf-8')
            
            # Obtener la clasificación de residuos que se va a editar
            clasificacion = get_object_or_404(CLASIFICACION_RESIDUOS, id=classification_id)
            
            # Crear el formulario con los datos de la clasificación
            form = ClasificacionResiduosForm(instance=clasificacion)
            return render(request, self.template_name, {'form': form, 'clasificacion': clasificacion})
        except Exception as e:
            print(e)
            return HttpResponseBadRequest('Error interno del servidor')

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Decodificar el ID en base64 dos veces
            class_key = kwargs.get('pk')
            classification_id = base64.urlsafe_b64decode(base64.urlsafe_b64decode(class_key)).decode('utf-8')
            
            # Obtener la clasificación de residuos que se va a editar
            clasificacion = get_object_or_404(CLASIFICACION_RESIDUOS, id=classification_id)
            
            form = ClasificacionResiduosForm(request.POST, request.FILES, instance=clasificacion)
            if form.is_valid():
                # Asignar el usuario actual a last_updated_by
                form.instance.last_updated_by = request.user
                # Guardar los cambios si el formulario es válido
                clasificacion = form.save()

                # Registrar evento
                tipo_evento = 'EDICION CLASIFICACION DE RESIDUOS'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = 'Clasificación de residuos actualizada correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                return JsonResponse({'success': False, 'errors': form.errors})
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor:{mensaje}')

# ------------------------------------ #
# Desactivar Clasificación de Residuos #

class DisableWasteSorting(View):
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            class_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(class_key).decode('utf-8')
            classification_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia de la clasificación de residuos
            clasificacion = get_object_or_404(CLASIFICACION_RESIDUOS, id=classification_id)
            # incluir el usuario que realiza el cambio
            clasificacion.last_updated_by= request.user
            # Desactivar la clasificación de residuos
            clasificacion.is_active = False
            clasificacion.save()
            
            # Registrar evento
            tipo_evento = 'DESACTIVAR CLASIFICACION DE RESIDUOS'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Clasificación "{clasificacion.name}" desactivada correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})

# ------------------------------------------------- #        
# Eliminar Registro de residuos de la base de datos #
class DeleteWasteRecordView(LoginRequiredMixin, View):

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            record_key = kwargs.get('pk')
            record_id_encoded = base64.urlsafe_b64decode(record_key).decode('utf-8')
            record_id = base64.urlsafe_b64decode(record_id_encoded).decode('utf-8')
            
            # Obtener la instancia del registro de residuos
            registro_residuo = get_object_or_404(REGISTRO_RESIDUOS, id=record_id)
            
                        
            # Eliminar el registro de residuos
            registro_residuo.delete()
            
            # Registrar evento
            tipo_evento = 'ELIMINAR REGISTRO DE RESIDUOS'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)
            
            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Registro de residuos "{registro_residuo.nombre_residuo}" eliminado correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})

        
# --------------------------------- #
# Activar Clasificación de Residuos #

class EnableWasteSorting(View):
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            class_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(class_key).decode('utf-8')
            classification_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia de la clasificación de residuos
            clasificacion = get_object_or_404(CLASIFICACION_RESIDUOS, id=classification_id)
            # incluir el usuario que realiza el cambio
            clasificacion.last_updated_by= request.user
            # Desactivar la clasificación de residuos
            clasificacion.is_active = True
            clasificacion.save()

            # Registrar evento
            tipo_evento = 'ACTIVAR CLASIFICACION DE RESIDUOS'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)
            
            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Clasificación "{clasificacion.name}" activada correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})
        
# -------------------------- #
# Crear Registro de Residuos #

class CreateRegisterWaste(LoginRequiredMixin, View):
    template_name = 'UniCLab_Residuos/crear_registro_residuos.html'

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL', 'COORDINADOR', 'TECNICO'])
    def get(self, request, *args, **kwargs):
        form = RegistroResiduosForm()
        laboratorios=Laboratorios.objects.all()
        unidades=Unidades.objects.all()
        
        lab_id=request.user.lab.id
        

        return render(request, self.template_name, {'form': form, 'laboratorios':laboratorios, 'lab_id':lab_id, 'unidades':unidades,})

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL', 'COORDINADOR', 'TECNICO'])
    def post(self, request, *args, **kwargs):
        form = RegistroResiduosForm(request.POST, request.FILES)
        try:
            if form.is_valid():
                # Asignar el usuario actual a los campos created_by y last_updated_by
                form.instance.created_by = request.user
                form.instance.last_updated_by = request.user

                
                # Guardar el registro si el formulario es válido
                registro_residuo = form.save()
                registro_residuo.total_residuo=registro_residuo.cantidad*registro_residuo.numero_envases
                registro_residuo.save()

                                
                # Registrar evento
                tipo_evento = 'CREAR REGISTRO DE RESIDUOS'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = 'Residuo agregado correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                print(form.errors)
                return JsonResponse({'success': False, 'errors': form.errors})
                
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor: {mensaje}')
        
# ------------------------------------------- #
# Listado de Registro de residuos de residuos #
class Wastes_Record_List(LoginRequiredMixin,ListView):
    model = REGISTRO_RESIDUOS
    template_name = "UniCLab_Residuos/listado_registros_residuos.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR','ADMINISTRADOR AMBIENTAL','COORDINADOR','TECNICO'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        
        # Obtener las fechas de inicio y fin de la solicitud GET
        id_laboratorio = self.request.GET.get('id_laboratorio')
        name = self.request.GET.get('name')
        sol = self.request.GET.get('sol')
        start_date = self.request.GET.get('start_date')   
        end_date = self.request.GET.get('end_date')   
        # Guardar los valores de filtrado en la sesión
        
        request.session['filtered_id_laboratorio'] = id_laboratorio
        request.session['filtered_name'] = name
        request.session['filtered_sol'] = sol
        request.session['filtered_start_date'] = start_date
        request.session['filtered_end_date'] = end_date
        
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        
        # Otros datos que quieras pasar al contexto
        context['usuarios'] = User.objects.all()
        context['laboratorios'] = Laboratorios.objects.all()
        # Obtener la fecha de hoy
        today = timezone.now().date()

        # Pasar la fecha de hoy al contexto
        context['today'] = today

        # Obtener el máximo valor de registro_solicitud donde residuo_enviado=True en REGISTRO_RESIDUOS
        max_solicitud = REGISTRO_RESIDUOS.objects.filter(residuo_enviado=True).aggregate(Max('registro_solicitud'))['registro_solicitud__max']
        min_solicitud = REGISTRO_RESIDUOS.objects.filter(residuo_enviado=True).aggregate(Min('registro_solicitud'))['registro_solicitud__min']
        # Pasar el máximo valor de registro_solicitud al contexto
        context['max_solicitud'] = max_solicitud
        context['min_solicitud'] = min_solicitud

        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()

        id_laboratorio = self.request.GET.get('id_laboratorio')

        name = self.request.GET.get('name')
        sol = self.request.GET.get('sol')
        
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # Validar y convertir las fechas
        try:
            if start_date:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            if end_date:
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
            pass

        
        # Realiza la filtración de acuerdo a las fechas
        if start_date:
            queryset = queryset.filter(date_create__gte=start_date, residuo_enviado=True)
        if end_date:
            queryset = queryset.filter(date_create__lte=end_date, residuo_enviado=True)
        elif start_date and end_date:
            queryset = queryset.filter(date_create__gte=start_date, date_create__lte=end_date, residuo_enviado=True)


        # Filtrar por  campos que contengan la palabra clave en su nombre
        if name:
            queryset = queryset.filter((Q(nombre_residuo__icontains=name) | Q(area__icontains=name) | Q(laboratorio__name__icontains=name) | Q(clasificado__name__icontains=name)), residuo_enviado=True)
        
        if sol:
            queryset=queryset.filter(registro_solicitud=sol,  residuo_enviado=True)
        
        if self.request.user.rol.name == 'ADMINISTRADOR' or self.request.user.rol.name == 'ADMINISTRADOR AMBIENTAL':
            
            if id_laboratorio:
                queryset = queryset.filter((Q(created_by__lab=id_laboratorio) | Q(laboratorio=id_laboratorio)), residuo_enviado=True)
        else:
            
            queryset = queryset.filter((Q(created_by__lab=self.request.user.lab) | Q(laboratorio=self.request.user.lab)), residuo_enviado=True)
        
        queryset = queryset.order_by('-registro_solicitud__id')
        return queryset
    
# --------------------------------------------- #
# Paginación de listado de registro de residuos #
class Wastes_Record_Pagination(LoginRequiredMixin,View):
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        filtered_id_laboratorio = request.session.get('filtered_id_laboratorio')
        filtered_name = request.session.get('filtered_name')
        filtered_sol = request.session.get('filtered_sol')
        filtered_start_date = request.session.get('filtered_start_date')
        filtered_end_date = request.session.get('filtered_end_date')
        
        url = reverse('residuos:record_waste_list')
        params = {}
        
        if filtered_id_laboratorio:
            params['id_laboratorio'] = filtered_id_laboratorio
        if filtered_name:
            params['name'] = filtered_name
        if filtered_sol:
            params['sol'] = filtered_sol
        if filtered_start_date:
            params['start_date'] = filtered_start_date
        if filtered_end_date:
            params['end_date'] = filtered_end_date
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)

    
# --------------------------------- #
# Autocompletar listado de residuos #
class AutocompleteWaste(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'COORDINADOR', 'TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request):
        term = request.GET.get('term', '')

        # Filtrar los residuos y obtener valores únicos
        residuos = REGISTRO_RESIDUOS.objects.filter(
            nombre_residuo__icontains=term
        ).values('nombre_residuo').distinct()[:10]

        # Convertir a mayúsculas y obtener valores únicos
        nombres_mayuscula = set(residuo['nombre_residuo'].upper() for residuo in residuos)

        # Crear una lista de resultados
        results = [{'name': nombre} for nombre in nombres_mayuscula]

        return JsonResponse(results, safe=False)
    
# ------------------------------------- #
# Exportar a Excel registro de residuos #
class Export2ExcelWastes(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
       
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        def get_file_name():
            now = datetime.now()
            return f'Registro_de_residuos_{now.strftime("%d%m%Y%H%M%S")}.xlsx'
           
        queryset=REGISTRO_RESIDUOS.objects.all()
        
        # Obtener datos filtrados de la sesión
        id_laboratorio = request.session.get('filtered_id_laboratorio')
        name = request.session.get('filtered_name')
        sol = request.session.get('filtered_sol')
        start_date = request.session.get('filtered_start_date')
        end_date = request.session.get('filtered_end_date')
        
        # Validar y convertir las fechas
        try:
            if start_date:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            if end_date:
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
            pass

        
        # Realiza la filtración de acuerdo a las fechas
        if start_date:
            queryset = queryset.filter(date_create__gte=start_date, residuo_enviado=True)
        if end_date:
            queryset = queryset.filter(date_create__lte=end_date, residuo_enviado=True)
        elif start_date and end_date:
            queryset = queryset.filter(date_create__gte=start_date, date_create__lte=end_date, residuo_enviado=True)


        # Filtrar por  campos que contengan la palabra clave en su nombre
        if name:
            queryset = queryset.filter((Q(nombre_residuo__icontains=name) | Q(area__icontains=name) | Q(laboratorio__name__icontains=name) | Q(clasificado__name__icontains=name)), residuo_enviado=True)
        
        if sol:
            queryset=queryset.filter(registro_solicitud=sol,  residuo_enviado=True)
        
        if self.request.user.rol.name == 'ADMINISTRADOR' or self.request.user.rol.name == 'ADMINISTRADOR AMBIENTAL':
            
            if id_laboratorio:
                queryset = queryset.filter((Q(created_by__lab=id_laboratorio) | Q(laboratorio=id_laboratorio)), residuo_enviado=True)
        else:
            
            queryset = queryset.filter((Q(created_by__lab=self.request.user.lab) | Q(laboratorio=self.request.user.lab)), residuo_enviado=True)
        
        queryset = queryset.order_by('registro_solicitud')

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')
        pil_image = PILImage.open(logo_path)
        image = ExcelImage(pil_image)
        sheet.add_image(image, 'A1')

        fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        sheet['C1'] = 'Listado de registro de residuos'
        sheet['C2'] = 'Fecha de Creación: '+fecha_creacion
        sheet['A4'] = 'ID Solicitud'
        sheet['B4'] = 'Dependencia'
        sheet['C4'] = 'Nombre del residuo'
        sheet['D4'] = 'Cantidad'
        sheet['E4'] = 'Número de envases'
        sheet['F4'] = 'Total Residuo'
        sheet['G4'] = 'Unidades'
        sheet['H4'] = 'Clasificado Y - A'
        sheet['I4'] = 'Estado'
        sheet['J4'] = 'Dependencia que realiza registro'
        sheet['K4'] = 'Usuario que registra'
        

        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 30
        sheet.row_dimensions[3].height = 25

        Titulo = sheet['C1']
        Titulo.font = Font(bold=True, size=16)
        Titulo.alignment = Alignment(horizontal='left')

        Fecha = sheet['C2']
        Fecha.alignment = Alignment(horizontal='left')
        
        

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        bold_font = Font(bold=True)

        # Establecer Anchos personalizados
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 31
        sheet.column_dimensions['C'].width = 27
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 21
        sheet.column_dimensions['K'].width = 31
        sheet.column_dimensions['L'].width = 10

        alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        

        row = 4
        for col in range(1, 13):
            sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).font = bold_font
            sheet.cell(row=row, column=col).alignment = alignment

        row = 5
        for item in queryset:
            # Crear la dependencia

            if item.laboratorio and item.area:
                dependencia=f'{item.laboratorio.name} {item.area}'
            elif item.laboratorio:
                dependencia=f'{item.laboratorio.name}'
            elif item.area:
                dependencia=f'{item.area}'
            
            # Organizar las clasificaciones
            clasificaciones = ', '.join(item.clasificado.values_list('name', flat=True)) if item.clasificado.exists() else 'NO DEFINIDA'

            
            sheet.cell(row=row, column=1).value = (item.registro_solicitud.id)
            sheet.cell(row=row, column=2).value = (dependencia)
            sheet.cell(row=row, column=3).value = item.nombre_residuo
            sheet.cell(row=row, column=4).value = item.cantidad
            sheet.cell(row=row, column=5).value = item.numero_envases
            sheet.cell(row=row, column=6).value = item.total_residuo
            sheet.cell(row=row, column=7).value = item.unidades.name
            sheet.cell(row=row, column=8).value = clasificaciones
            sheet.cell(row=row, column=9).value = item.estado.name
            sheet.cell(row=row, column=10).value =item.created_by.lab.name
            sheet.cell(row=row, column=11).value = f'{item.created_by.first_name} {item.created_by.last_name}'
              
            
            for col in range(1, 13):
                sheet.cell(row=row, column=col).border = thin_border
                sheet.cell(row=row, column=col).alignment = alignment

            row += 1

        start_column = 1
        end_column = 12
        start_row = 4
        end_row = row - 1

        start_column_letter = get_column_letter(start_column)
        end_column_letter = get_column_letter(end_column)

        table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

        sheet.auto_filter.ref = table_range

        fill = PatternFill(fill_type="solid", fgColor=WHITE)
        start_cell = sheet['A1']
        end_column_letter = get_column_letter(end_column+1)
        end_row = row+1
        end_cell = sheet[end_column_letter + str(end_row)]
        table_range = start_cell.coordinate + ':' + end_cell.coordinate

        for row in sheet[table_range]:
            for cell in row:
                cell.fill = fill

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={get_file_name()}'

        workbook.save(response)

        tipo_evento = 'DESCARGA DE ARCHIVOS'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        return response
    
# --------------------------- #
# Editar Registro de Residuos #
class EditarRegistroResiduos(LoginRequiredMixin, View):
    template_name = 'UniCLab_Residuos/editar_registro_residuos.html'

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        try:
            # Decodificar el ID en base64 dos veces
            registro_key = kwargs.get('pk')
            registro_id = base64.urlsafe_b64decode(base64.urlsafe_b64decode(registro_key)).decode('utf-8')
            # Obtener el registro de residuos que se va a editar
            registro_residuo = get_object_or_404(REGISTRO_RESIDUOS, id=registro_id)
            
            # Crear el formulario con los datos del registro
            form = EditResiduosForm(instance=registro_residuo)
            laboratorios = Laboratorios.objects.all()
            lab_id = request.user.lab.id
            unidades = Unidades.objects.all()

            return render(request, self.template_name, {'form': form, 'laboratorios': laboratorios, 'lab_id': lab_id,'residuo': registro_residuo,'unidades':unidades,})
        except Exception as e:
            print(e)
            return HttpResponseBadRequest('Error interno del servidor')

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Decodificar el ID en base64 dos veces
            registro_key = kwargs.get('pk')
            registro_id = base64.urlsafe_b64decode(base64.urlsafe_b64decode(registro_key)).decode('utf-8')
            
            
            # Obtener el registro de residuos que se va a editar
            registro_residuo = get_object_or_404(REGISTRO_RESIDUOS, id=registro_id)
            
            form = EditResiduosForm(request.POST, request.FILES, instance=registro_residuo)
            if form.is_valid():
                # Asignar el usuario actual a last_updated_by
                form.instance.last_updated_by = request.user
                # Totalizar el residuo
                registro_residuo.total_residuo=registro_residuo.cantidad*registro_residuo.numero_envases
                # Guardar los cambios si el formulario es válido
                registro_residuo = form.save()

                # Registrar evento
                tipo_evento = 'EDICION REGISTRO DE RESIDUOS'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = 'Registro de residuos actualizado correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                return JsonResponse({'success': False, 'errors': form.errors})
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor:{mensaje}')


# ------------------------------------ #
# Desactivar Clasificación de Residuos #

class DisableWasteRecord(View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            waste_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(waste_key).decode('utf-8')
            waste_record_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia de la clasificación de residuos
            wate_record = get_object_or_404(REGISTRO_RESIDUOS, id=waste_record_id)
            # incluir el usuario que realiza el cambio
            wate_record.last_updated_by= request.user
            # Desactivar la clasificación de residuos
            wate_record.is_active = False
            wate_record.save()
            
            # Registrar evento
            tipo_evento = 'DESACTIVAR REGISTRO DE RESIDUOS'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Registro de residuo"{wate_record.nombre_residuo}" deshabilitado correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})
        
# --------------------------------- #
# Activar Clasificación de Residuos #

class EnableWasteRecord(View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            waste_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(waste_key).decode('utf-8')
            waste_record_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia de la clasificación de residuos
            wate_record = get_object_or_404(REGISTRO_RESIDUOS, id=waste_record_id)
            # incluir el usuario que realiza el cambio
            wate_record.last_updated_by= request.user
            # Desactivar la clasificación de residuos
            wate_record.is_active = True
            wate_record.save()
            
            # Registrar evento
            tipo_evento = 'ACTIVAR REGISTRO DE RESIDUOS'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Registro de residuo"{wate_record.nombre_residuo}" habilitado correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})
        
# ----------------------------- #        
# Vista para cambiar contraseña #        
class PasswordChangeView(PasswordContextMixin, FormView):
    form_class = PasswordChangeForm
    success_url = reverse_lazy("residuos:home_residuos")
    template_name = "UniCLab_Residuos/cambiar_pass.html"
    title = ("Password change")

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    @method_decorator(sensitive_post_parameters())
    @method_decorator(csrf_protect)
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.save()
        tipo_evento = 'CAMBIAR CONTRASENA'
        usuario_evento = self.request.user
        crear_evento(tipo_evento, usuario_evento)

        update_session_auth_hash(self.request, form.user)

        # Agregar un mensaje de éxito
        messages.success(self.request, 'Contraseña cambiada exitosamente.')

        return super().form_valid(form)

# ------------- #
# Cerrar Sesión #
class LogoutView(RedirectURLMixin, TemplateView):
    """
    Log out the user and display the 'You are logged out' message.
    """

    # RemovedInDjango50Warning: when the deprecation ends, remove "get" and
    # "head" from http_method_names.
    http_method_names = ["get", "head", "post", "options"]
    template_name = "UniCLab_Residuos/logged_out_residuos.html"
    extra_context = None

    # RemovedInDjango50Warning: when the deprecation ends, move
    # @method_decorator(csrf_protect) from post() to dispatch().
    @method_decorator(never_cache)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    @method_decorator(csrf_protect)
    def post(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            # Crea un evento de cierre de sesión
            tipo_evento = 'CIERRE DE SESION'
            usuario_evento = self.request.user
            crear_evento(tipo_evento, usuario_evento)

        """Logout may be done via POST."""
        auth_logout(request)
        redirect_to = self.get_success_url()
        if redirect_to != request.get_full_path():
            # Redirect to target page once the session has been cleared.
            return HttpResponseRedirect(redirect_to)
        return super().get(request, *args, **kwargs)

    # RemovedInDjango50Warning.
    get = post

    def get_default_redirect_url(self):
        """Return the default redirect URL."""
        if self.next_page:
            return resolve_url(self.next_page)
        elif settings.LOGOUT_REDIRECT_URL:
            return resolve_url(settings.LOGOUT_REDIRECT_URL)
        else:
            return self.request.path

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_site = get_current_site(self.request)
        context.update(
            {
                "site": current_site,
                "site_name": current_site.name,
                "title": ("Logged out"),
                "subtitle": None,
                **(self.extra_context or {}),
            }
        )
        return context
    
    def logout_then_login(request, login_url=None):
            """
            Log out the user if they are logged in. Then redirect to the login page.
            """
            login_url = resolve_url(login_url or settings.LOGIN_URL)
            return LogoutView.as_view(next_page=login_url)(request)
        
    def redirect_to_login(next, login_url=None, redirect_field_name=REDIRECT_FIELD_NAME):
            """
            Redirect the user to the login page, passing the given 'next' page.
            """
            resolved_url = resolve_url(login_url or settings.LOGIN_URL)
    
            login_url_parts = list(urlparse(resolved_url))
            if redirect_field_name:
                querystring = QueryDict(login_url_parts[4], mutable=True)
                querystring[redirect_field_name] = next
                login_url_parts[4] = querystring.urlencode(safe="/")
    
            return HttpResponseRedirect(urlunparse(login_url_parts))
    
# ------------------------------- #
# Listado de solicitudes Externas #
class SolicitudesExtListView(LoginRequiredMixin,ListView):
    model = SolicitudesExternas
    template_name = "UniCLab_Residuos/listado_solicitudes_externas.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR','ADMINISTRADOR AMBIENTAL'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        lab = self.request.GET.get('lab')
        keyword = self.request.GET.get('keyword')  
        
        

        # Guardar los valores de filtrado en la sesión
        
        request.session['filtered_start_date'] = start_date
        request.session['filtered_end_date'] = end_date
        request.session['filtered_lab'] = lab
        request.session['filtered_keyword'] = keyword
        

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener la fecha de hoy
        today = date.today()
        # Calcular la fecha hace un mes hacia atrás
        one_month_ago = today - timedelta(days=30)

        # Agregar la fecha al contexto
        context['one_month_ago'] = one_month_ago

        # Agregar la fecha de hoy al contexto
        context['today'] = today
        laboratorio = self.request.user.lab
        
    
        context['usuarios'] = User.objects.all()
        context['laboratorio'] = laboratorio
        context['laboratorios'] = Laboratorios.objects.all()
        
        # Obtener la lista de inventarios
        entradas = context['object_list']
        # Recorrer los entradas y cambiar el formato de la fecha        
               
        context['object_list'] = entradas
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()

        lab = self.request.GET.get('lab')
        keyword = self.request.GET.get('keyword')
        
        

        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
             lab=''

        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # Validar y convertir las fechas
        try:
            if start_date:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            if end_date:
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
            pass

        
        # Realiza la filtración de acuerdo a las fechas
        if start_date:
            queryset = queryset.filter(registration_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(registration_date__lte=end_date)
        elif start_date and end_date:
            queryset = queryset.filter(registration_date__gte=start_date,date_create__lte=end_date)
        
        # Filtrar por  campos que contengan la palabra clave en su nombre
        if keyword:
            queryset = queryset.filter(Q(name__icontains=keyword) | Q(subject__icontains=keyword) | Q(message__icontains=keyword) | Q(attach__icontains=keyword))
        
        if lab:
            queryset = queryset.filter(lab=lab)

        queryset = queryset.order_by('-id')
        return queryset
    
# ------------------------------------------- #
# Paginación de solicitudes externas residuos #
class GuardarPerPageViewExternalRequests(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        filtered_start_date = request.session.get('filtered_start_date')
        filtered_end_date = request.session.get('filtered_end_date')
        filtered_lab = request.session.get('filtered_lab')
        filtered_keyword = request.session.get('filtered_keyword')
        
        url = reverse('residuos:external_requests')
        params = {}
        
        if filtered_start_date:
            params['start_date'] = filtered_start_date
        if filtered_end_date:
            params['end_date'] = filtered_end_date
        if filtered_lab:
            params['lab'] = filtered_lab
        if filtered_keyword:
            params['keyword'] = filtered_keyword
        
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)
    

# ----------------------- #
# Registro de solicitudes #
class RegistrarSolicitudResiduos(LoginRequiredMixin, CreateView):
    model = Solicitudes
    form_class = SolicitudForm
    template_name = 'UniCLab_Residuos/registrar_solicitud.html'
    success_url = '/'

    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agregar el usuario en sesión y el laboratorio al contexto
        context['usuario'] = self.request.user
        context['laboratorio'] = self.request.user.lab
        return context
    # Decorador para asegurar que la función get tenga el permiso requerido
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)
    
    @check_group_permission(groups_required=['ADMINISTRADOR','ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)

        
        
        if form.is_valid():
            solicitud = form.save(commit=False)  # No guardar aún, solo crear una instancia
            solicitud.created_by = request.user
            solicitud.last_updated_by = request.user
            solicitud.save()  # guarda la instancia con los archivos
            form.save()
            # Crea un evento de registrar solicitud
            tipo_evento = 'REGISTRAR SOLICITUD'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            mensaje = f'La solicitud se ha registrado correctamente, se a enviado un correo al administrador del sistema el radicado de su solicitud es: {solicitud.id:04}'
            messages.success(request, mensaje)  # Agregar mensaje de éxito
            
            # Enviar correo al usuario que registra la solicitud
            id = solicitud.id
            solicitud_code = base64.urlsafe_b64encode(str(id).encode()).decode()
            suffix = f'UniCLab/solicitudes/estado_solicitud/{solicitud_code}'
            shipping_email=solicitud.created_by.email
            email_type=f'Registro de solicitud'
            initial_message= f'El presente mensaje de correo electrónico es porque recientemente se ha registrado una solicitud en el aplicativo de inventario y gestión de reactivos, con la siguiente información:'
            # enviar_correo_solicitud(request, suffix, id, initial_message,shipping_email, email_type )
            # envío de correo electrónico en segundo plano
        
            # Crear un hilo y ejecutar enviar_correo en segundo plano
            correo_thread = threading.Thread(
            target=enviar_correo_solicitud, args=(request, suffix, id, initial_message,shipping_email, email_type),)
            correo_thread.start()
            # Enviar correo al administrador del aplicativo
            # Obtener el correo del administrador desde la base de datos
            configuracion = ConfiguracionSistema.objects.first()
            if configuracion:
                admin_email = configuracion.correo_administrador
            else:
                admin_email = 'uniclab_man@unal.edu.co'

            id = solicitud.id
            protocol = 'https' if request.is_secure() else 'http'
            domain = request.get_host()
            url=f'{protocol}://{domain}/UniCLab/solicitudes/listado_solicitudes'
            solicitud_code = base64.urlsafe_b64encode(str(id).encode()).decode()
            suffix = f'UniCLab/solicitudes/responder_solicitud/{solicitud_code} o visita: {url}'
            shipping_email=admin_email
            email_type=f'Registro de solicitud'
            initial_message= f'Recientemente se ha registrado una solicitud en el aplicativo de inventario y gestión de reactivos, con la siguiente información:'
            # enviar_correo_solicitud(request, suffix, id, initial_message,shipping_email,email_type )
            # Crear un hilo y ejecutar enviar_correo en segundo plano
            correo_thread = threading.Thread(
            target=enviar_correo_solicitud, args=(request, suffix, id, initial_message,shipping_email, email_type),)
            correo_thread.start()
        else:
            mensaje = f'La solicitud no se ha podido enviar, por favor consulte el administrador del sistema, error:.'
            error = form.errors
            mensaje = f'{mensaje} {error}'
            messages.error(request, mensaje)  # Agregar mensaje de error

        return redirect('residuos:internal_requests_form')

# ------------------------------------------- #
# Vista para acumular Residuos en solicitud #
class Temporal_Wastes_Record(LoginRequiredMixin,ListView):
    model = REGISTRO_RESIDUOS
    template_name = "UniCLab_Residuos/solicitud_registros_residuos.html"
    paginate_by = 100
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR','ADMINISTRADOR AMBIENTAL','COORDINADOR','TECNICO'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        
        
        
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        
        # Otros datos que quieras pasar al contexto
        context['usuarios'] = User.objects.all()
        context['laboratorios'] = Laboratorios.objects.all()
        

        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        lab=self.request.user.lab

        queryset = queryset.filter(created_by__lab=lab, residuo_enviado=False)

        queryset = queryset.order_by('id')
        return queryset
    

# --------------------------------------- #
# Enviar Solicitud de residuo de Residuos #

class SendWasteRecord(LoginRequiredMixin, View):
    def enviar_correo_asincrono(self, recipient_list, subject, message, attach_path):
        try:
            enviar_correo(recipient_list, subject, message, attach_path)
        except Exception as e:
            print(f'Error al enviar correo: {e}')
    
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL', 'COORDINADOR', 'TECNICO'])
    def post(self, request, *args, **kwargs):
        try:
            registros = REGISTRO_RESIDUOS.objects.filter(
            created_by__lab=self.request.user.lab,
            residuo_enviado=False
            )
            if registros:
                # Lista para almacenar el contenido dinámico de cada registro afectado
                prebody = []
                configuracion=ConfiguracionSistema.objects.first()
                url=configuracion.url
            

                # Actualizar los registros y obtener los datos necesarios para el correo
                for registro in registros:
                    # Observaciones
                    if registro.observaciones:
                        observaciones=f"<b>Observaciones:</b> {registro.observaciones}<br>"
                    else:
                        observaciones=''

                    # Dependencia

                    if registro.laboratorio and registro.area:
                        dependencia=f'{registro.laboratorio.name}, {registro.area}'
                    elif registro.laboratorio:
                        dependencia=f'{registro.laboratorio.name}'
                    else:
                        dependencia=f'{registro.area}'


                    
                    

                    # Construir el contenido dinámico para cada registro
                    prebody.append(f"<b>Dependencia:</b> {dependencia}<br>"
                                    f"<b>Residuo:</b> {registro.nombre_residuo}<br>"
                                   f"<b>Cantidad:</b> {registro.cantidad} {registro.unidades}<br>"
                                   f"<b>Envases:</b> {registro.numero_envases}<br>"
                                   f"<b>Total residuo:</b> {registro.total_residuo} {registro.unidades}<br>"
                                   f"<b>Clasificación:</b> {', '.join(str(clasif) for clasif in registro.clasificado.all())}<br>"
                                   f"<b>Estado:</b> {registro.estado}<br>"
                                   f"{observaciones}"
                                   
                                   
                                   )
                    

                    
                # Crear nueva solicitud
                solicitud = SOLICITUD_RESIDUO.objects.create(created_by=request.user, last_updated_by=request.user)
                solicitud.name='Sol_' + str(solicitud.pk)
                solicitud.save()
                registros.update(residuo_enviado=True, registro_solicitud=solicitud)
                # Registrar evento
                tipo_evento = 'ENVIO REGISTRO DE RESIDUOS'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)
                # Datos para correo electrónico

                # Usuarios que recepcionan
                # Obtener el correo del usuario que realiza el registro
                correo_usuario = request.user.email

                # Obtener los correos de todos los usuarios activos con rol 'ADMINISTRADOR AMBIENTAL'
                usuarios_admin_ambiental = User.objects.filter(is_active=True, rol__name='ADMINISTRADOR AMBIENTAL').values_list('email', flat=True)

                # Crear una lista de destinatarios
                recipient_list = list(usuarios_admin_ambiental)
                recipient_list.append(correo_usuario)  # Agregar el correo del usuario que realiza el registro

                # Asunto
                subject= f'Registro de residuos en UniCLab Residuos - Consecutivo: {solicitud.pk}'

                # Mensaje
                # Definir como dependencia el area o laboratorio
                
                # Formatear fecha    
                # Obtener el huso horario deseado (GMT-5)
                tz = pytz.timezone('America/Bogota')

                # Convertir la fecha al huso horario deseado y formatearla
                fecha_registro = solicitud.date_create.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
                
    
                
                print(prebody)
                header=f'<p>El siguiente mensaje tiene el fin de informar que recientemente se ha realizado un registro de residuos dirgido a la <i>OFICINA DE GESTIÓN AMBIENTAL</i> de la Universidad Nacional de Colombia, Sede Manizales. Los datos del registro son los siguientes:</p><p><b>Fecha del registro: </b>{fecha_registro}<br><b>Consecutivo: </b>{solicitud.pk}<br><b>Responsable del registro: </b>{solicitud.created_by.first_name} {solicitud.created_by.last_name}<br><b>Correo electrónio: </b>{solicitud.created_by.email}<br><b>Dependencia que registra: </b>{solicitud.created_by.lab}</p>'
                body = '<br>'.join(prebody)
                footer=f'<p>Este correo es informativo, para más detalle dirígete a la web principal de UniCLab Residuos.</p>'
                message=header+body+footer
                
                # Adjuntos
                attach_path=None
               
                
                # Crear un hilo y ejecutar enviar_correo en segundo plano
                correo_thread = threading.Thread(
                    target=self.enviar_correo_asincrono,
                    args=(recipient_list, subject, message, attach_path),
                )
                correo_thread.start()

                # Devolver una respuesta JSON de éxito
                return JsonResponse({'success': True, 'message': f'Se han registrado los residuos correctamente.'})
            else:
                return JsonResponse({'success': False, 'message': f'Debe agregar residuos primero.'})

        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})
        

# --------------------------------- #
# Eliminar Residuo de la lista 'temporal' #

class DeleteWaste(LoginRequiredMixin, View):
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL', 'COORDINADOR', 'TECNICO'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            waste_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(waste_key).decode('utf-8')
            waste_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia de la clasificación de residuos
            residuo = get_object_or_404(REGISTRO_RESIDUOS, id=waste_id)
            
            
            
            # Eliminar el registro de residuos
            residuo.delete()
           
            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Residuo eliminado de la lista.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})


# --------------------------------------- #
# Cancelar Solicitud de residuo de Residuos #

class CancelWasteRecord(LoginRequiredMixin, View):
   
    
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL', 'COORDINADOR', 'TECNICO'])
    def post(self, request, *args, **kwargs):
        try:
            registros = REGISTRO_RESIDUOS.objects.filter(
            created_by__lab=self.request.user.lab,
            residuo_enviado=False
            )
            if registros:
                

                
                # Registrar evento
                tipo_evento = 'CANCELAR ENVIO REGISTRO DE RESIDUOS'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)
                

                # Devolver una respuesta JSON de éxito
                return JsonResponse({'success': True, 'message': f'Se ha cancelado la solicitud'})
            else:
                return JsonResponse({'success': False, 'message': f'No tiene registros para cancelar envío.'})

        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})
        
    


# ---------------------------------- #
# Ver detalle de la solicitud en PDF #
class SolicitudPDFEmbView(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL', 'COORDINADOR', 'TECNICO'])
    def get(self, request, pk):
        
        # Obtener la solicitud por su id (pk)
        # Obtener el ID codificado dos veces desde los parámetros de la solicitud
        request_key = pk
        item_id_encoded = base64.urlsafe_b64decode(request_key).decode('utf-8')
        request_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
        solicitud = get_object_or_404(SOLICITUD_RESIDUO, pk=request_id)
        
        # Obtener los registros de residuos asociados a la solicitud
        registros = REGISTRO_RESIDUOS.objects.filter(registro_solicitud=solicitud)
        # Generar el PDF y obtener el nombre del archivo
        logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')
        
        
        # Construir el contexto para la plantilla
        context = {
            'solicitud': solicitud,
            'registros': registros,
            'configuracion': ConfiguracionSistema.objects.first(),
            'logo_path':logo_path,
            
        }
        
        pdf_filename = render_to_pdf_file('UniCLab_Residuos/solicitudes_residuos.html', context)

        pdf_url = settings.MEDIA_URL + pdf_filename
        
        # Renderizar la plantilla HTML con el PDF embebido
        return render(request, 'UniCLab_Residuos/detalle_solicitud_residuos.html', {'pdf_url': pdf_url,'pdf_filename': pdf_filename,'solicitud': solicitud.id ,'registro': solicitud})

    
# ---------------------------------------------------- #
# Borrar Solicitud en PDF y marca solicitud como leída #

class DeletePDFView(LoginRequiredMixin, View):
    def enviar_correo_asincrono(self, recipient_list, subject, message, attach_path):
        try:
            enviar_correo(recipient_list, subject, message, attach_path)
            # Eliminar el archivo PDF después de enviar el correo electrónico
            if attach_path and os.path.exists(attach_path):
                os.remove(attach_path)
        except Exception as e:
            print(f'Error al enviar correo: {e}')

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL', 'COORDINADOR', 'TECNICO'])
    def get(self, request, *args, **kwargs):
        pdf_filename = request.GET.get('pdf_filename')
        solicitud_id = request.GET.get('solicitud')
        notification = request.GET.get('notification')
        days = request.GET.get('days')
        
        solicitud_registro = get_object_or_404(SOLICITUD_RESIDUO, id=solicitud_id)
        # Marcar la solicitud como leída si el usuario es administrador o administrador ambiental
        if request.user.rol.name=="ADMINISTRADOR" or request.user.rol.name=="ADMINISTRADOR AMBIENTAL":
            solicitud_registro.leido = True
            solicitud_registro.save()
        # Registrar evento
        tipo_evento = 'LECTURA DE SOLICITUD DE RESIDUOS'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        
        if notification == 'True':
            if request.user.rol.name=="ADMINISTRADOR" or request.user.rol.name=="ADMINISTRADOR AMBIENTAL":
                # Marcar la solicitud como respondida
                solicitud_registro.respondido = True
                solicitud_registro.save()
                # Registrar evento
                tipo_evento = 'RESPUESTA A SOLICITUD DE RESIDUOS'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)
                # Datos para correo electrónico

                # Usuarios que recepcionan
                # Obtener el correo del usuario que realiza el registro
                correo_usuario = request.user.email

                # Obtener los correos de todos los usuarios activos con rol 'ADMINISTRADOR AMBIENTAL'
                usuario_solicitud = solicitud_registro.created_by.email

                # Crear una lista de destinatarios
                recipient_list = [correo_usuario,usuario_solicitud]

                # Asunto
                subject= f'Respuesta a registro de residuos en UniCLab Residuos - Consecutivo: {solicitud_registro.pk}'

                # Mensaje

                # Formatear fecha    
                # Obtener el huso horario deseado (GMT-5)
                tz = pytz.timezone('America/Bogota')

                # Convertir la fecha al huso horario deseado y formatearla
                fecha_registro = solicitud_registro.date_create.astimezone(tz).strftime('%d/%m/%Y')
                hora_registro = solicitud_registro.date_create.astimezone(tz).strftime('%H:%M:%S')
                usuario_registro= f'{solicitud_registro.created_by.first_name} {solicitud_registro.created_by.last_name}'
                dependencia=f'{solicitud_registro.created_by.lab}'
                header=f'<p>Señor(a):<br><b><i>{usuario_registro}</i></b><br>{dependencia}<br><br>¡Cordial Saludo!</p>'
                body = f'<p>El presente mensaje tiene como finalidad informar que hemos revisado el registro realizado por usted el pasado {fecha_registro} a las {hora_registro}, con consecutivo # <b>{solicitud_registro.pk}</b>.<br><br>A partir de los próximos {days} días hábiles, tendrá respuesta a su registro.</p>'
                footer=f'<p>Este correo es informativo, para más detalle dirígete a la web principal de UniCLab Residuos.</p>'
                message=header+body+footer

                # Adjuntos
                attach_path = os.path.join(settings.MEDIA_ROOT, pdf_filename)

                # Crear un hilo y ejecutar enviar_correo en segundo plano
                correo_thread = threading.Thread(
                    target=self.enviar_correo_asincrono,
                    args=(recipient_list, subject, message, attach_path),
                )
                correo_thread.start()
                return JsonResponse({'status': 'success'})
        else:
            # Eliminar el archivo PDF si existe
            if pdf_filename:
                pdf_path = os.path.join(settings.MEDIA_ROOT, pdf_filename)
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                    return JsonResponse({'status': 'success'})
        
        return JsonResponse({'status': 'success'})

# ------------------------------------------ #
# Marcar Registro de  Residuos como no leido #

class markAsUnread(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            waste_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(waste_key).decode('utf-8')
            waste_record_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia de la clasificación de residuos
            wate_record = get_object_or_404(SOLICITUD_RESIDUO, id=waste_record_id)
            # incluir el usuario que realiza el cambio
            wate_record.last_updated_by= request.user
            # Desactivar la clasificación de residuos
            wate_record.leido = False
            wate_record.save()
            
            # Registrar evento
            tipo_evento = 'MARCAR REGISTRO COMO NO LEIDO'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Registro de residuo"{wate_record.id}" Se ha marcado como no leída.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})


# -------------------------------------------------- #
# Enviar notificación de lectura a solicitud externa #

# Enviar notificación de lectura a solicitud externa
class SendReadNotificationRequest(LoginRequiredMixin, View):
    def enviar_correo_asincrono(self, recipient_list, subject, message, attach_path):
        try:
            enviar_correo(recipient_list, subject, message, attach_path)
            # Eliminar el archivo PDF después de enviar el correo electrónico
            if attach_path and os.path.exists(attach_path):
                os.remove(attach_path)
        except Exception as e:
            print(f'Error al enviar correo: {e}')

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        dias = request.GET.get('dias')
        id_solicitud = request.GET.get('encoded_id')

        # Decodificar 2 veces el id de la solicitud
        request_id = base64.urlsafe_b64decode(id_solicitud).decode('utf-8')
        # Obtener la instancia de la solicitud externa
        solicitud = get_object_or_404(SolicitudesExternas, id=request_id)

        # Respuesta solo si es administrador o administrador ambiental
        if request.user.rol.name == "ADMINISTRADOR" or request.user.rol.name == "ADMINISTRADOR AMBIENTAL":
            # Marcar solicitud como respondida
            solicitud.has_answered = True
            solicitud.save()

            # Crear evento
            tipo_evento = 'RESPUESTA A SOLICITUD EXTERNA'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Datos para correo electrónico
            # Usuarios que recepcionan
            # Obtener el correo del usuario que responde
            correo_usuario = request.user.email

            # Obtener los correos del usuario que realiza la solicitud
            usuario_solicitud = solicitud.email
            # Crear una lista de destinatarios
            recipient_list = [correo_usuario, usuario_solicitud]
            # Asunto
            subject = f'Notificación de lectura a solicitud externa - {solicitud.subject}'
            # Mensaje
            # Formatear fecha
            # Obtener el huso horario deseado (GMT-5)
            tz = pytz.timezone('America/Bogota')
            # Convertir la fecha al huso horario deseado y formatearla
            fecha_registro = solicitud.registration_date.astimezone(tz).strftime('%d/%m/%Y')
            hora_registro = solicitud.registration_date.astimezone(tz).strftime('%H:%M:%S')
            usuario_registro = f'{solicitud.name}'
            dependencia = f'{solicitud.department}'
            header = f'<p>Señor(a):<br><b><i>{usuario_registro}</i></b><br>{dependencia}<br><br>¡Cordial Saludo!</p>'
            body = f'<p>El presente mensaje tiene como finalidad informar que hemos revisado el registro realizado por usted el pasado {fecha_registro} a las {hora_registro}, con consecutivo # <b>{solicitud.pk}</b>.<br><br>A partir de los próximos {dias} días hábiles, tendrá respuesta a la solicitud realizada.</p>'
            footer = f'<p>Este correo es informativo, para más detalle dirígete a la web principal de UniCLab Residuos.</p>'
            message = header + body + footer
            # Adjuntos
            attach_path = None
            # Crear un hilo y ejecutar enviar_correo en segundo plano
            correo_thread = threading.Thread(
                target=self.enviar_correo_asincrono,
                args=(recipient_list, subject, message, attach_path),
            )
            correo_thread.start()
            return JsonResponse({'status': 'success'})
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'})

# ------------------------- #
# Crear Fichas de Seguridad #

class CreateSecuritySheet(LoginRequiredMixin, View):
    template_name = 'UniCLab_Residuos/crear_ficha_seguridad.html'

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        form = FichaSeguridadForm()
        

        return render(request, self.template_name, {'form': form,})

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        form = FichaSeguridadForm(request.POST, request.FILES)
        try:
            if form.is_valid():
                # Asignar el usuario actual a los campos created_by y last_updated_by
                form.instance.created_by = request.user
                form.instance.last_updated_by = request.user

                
                # Guardar el registro si el formulario es válido
                registro_ficha = form.save()
                                                
                # Registrar evento
                tipo_evento = 'CREAR FICHA DE SEGURIDAD'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = f'Ficha de seguiridad {registro_ficha.name} creada correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                print(form.errors)
                return JsonResponse({'success': False, 'errors': form.errors})
                
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor: {mensaje}')

# ------------------------------------ #
# Visualización de Fichas de Seguridad #
class Security_Sheet_List(LoginRequiredMixin,ListView):
    model = FichaSeguridad
    template_name = "UniCLab_Residuos/listado_fichas_seguridad.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR','ADMINISTRADOR AMBIENTAL','COORDINADOR','TECNICO'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 100  # Valor predeterminado si no hay variable de sesión

        
        
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        
        # Otros datos que quieras pasar al contexto
        context['usuarios'] = User.objects.all()
        context['laboratorios'] = Laboratorios.objects.all()
        

        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()

        
        queryset = queryset.order_by('id')
        return queryset


# -------------------------- #
# Editar Fichas de Seguirdad #
class EditarSecuritySheet(LoginRequiredMixin, View):
    template_name = 'UniCLab_Residuos/editar_ficha_seguridad.html'

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        try:
            # Decodificar el ID en base64 dos veces
            ficha_key = kwargs.get('pk')
            ficha_id = base64.urlsafe_b64decode(base64.urlsafe_b64decode(ficha_key)).decode('utf-8')
            # Obtener el registro de residuos que se va a editar
            ficha_seguridad = get_object_or_404(FichaSeguridad, id=ficha_id)
            
            # Crear el formulario con los datos del registro
            form = FichaSeguridadForm(instance=ficha_seguridad)
            

            return render(request, self.template_name, {'form': form,})
        except Exception as e:
            print(e)
            return HttpResponseBadRequest('Error interno del servidor')

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Decodificar el ID en base64 dos veces
            ficha_key = kwargs.get('pk')
            ficha_id = base64.urlsafe_b64decode(base64.urlsafe_b64decode(ficha_key)).decode('utf-8')
            
            
            # Obtener el registro de residuos que se va a editar
            ficha_seguridad = get_object_or_404(FichaSeguridad, id=ficha_id)
            
            form = FichaSeguridadForm(request.POST, request.FILES, instance=ficha_seguridad)
            if form.is_valid():
                # Asignar el usuario actual a last_updated_by
                form.instance.last_updated_by = request.user
                
                ficha_seguridad = form.save()

                # Registrar evento
                tipo_evento = 'EDICION FICHA DE SEGUIRDAD'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = 'Ficha de seguirdad actualizado correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                return JsonResponse({'success': False, 'errors': form.errors})
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor:{mensaje}')

# ----------------------------- #
# Desactivar Ficha de seguridad #
class DisableSecuritySheet(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            ficha_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(ficha_key).decode('utf-8')
            ficha_seguridad_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia del registro
            ficha_seguridad = get_object_or_404(FichaSeguridad, id=ficha_seguridad_id)
            # incluir el usuario que realiza el cambio
            ficha_seguridad.last_updated_by= request.user
            # Desactivar el registro
            ficha_seguridad.is_active = False
            ficha_seguridad.save()
            
            # Registrar evento
            tipo_evento = 'DESACTIVAR FICHA DE SEGURIDAD'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Ficha de seguridad {ficha_seguridad.name} deshabilitada correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})
        
# -------------------------- #
# Activar Ficha de seguridad #
class EnableSecuritySheet(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            ficha_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(ficha_key).decode('utf-8')
            ficha_seguridad_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia del registro
            ficha_seguridad = get_object_or_404(FichaSeguridad, id=ficha_seguridad_id)
            # incluir el usuario que realiza el cambio
            ficha_seguridad.last_updated_by= request.user
            # Desactivar el registro
            ficha_seguridad.is_active = True
            ficha_seguridad.save()
            
            # Registrar evento
            tipo_evento = 'ACTIVAR FICHA DE SEGURIDAD'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Ficha de seguridad {ficha_seguridad.name} habilitada correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})
        

# ------------------------------------- #
# Crear Certificado de Disposción Final #

class FinalCertificate(LoginRequiredMixin, View):
    template_name = 'UniCLab_Residuos/crear_certificado_disposicion.html'

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        form = CertificadoDisposicionForm()
        

        return render(request, self.template_name, {'form': form,})

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        form = CertificadoDisposicionForm(request.POST, request.FILES)
        try:
            if form.is_valid():
                # Asignar el usuario actual a los campos created_by y last_updated_by
                form.instance.created_by = request.user
                form.instance.last_updated_by = request.user

                
                # Guardar el registro si el formulario es válido
                registro_ficha = form.save()
                                                
                # Registrar evento
                tipo_evento = 'CREAR CERTIFICADO DE DISPOSICION'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = f'Certificado de disposición creado correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                print(form.errors)
                return JsonResponse({'success': False, 'errors': form.errors})
                
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor: {mensaje}')
        
# ------------------------------------ #
# Visualización de Certificados de Disposición final #
class FinalCertificateList(LoginRequiredMixin,ListView):

    model = CERTIFICADO_DISPOSICION
    template_name = "UniCLab_Residuos/listado_certificados_disposicion.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        name = self.request.GET.get('name')
        # Guardar los valores de filtrado en la sesión
        request.session['filtered_name'] = name
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['usuarios'] = User.objects.all()
        
       
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        name = self.request.GET.get('name')
        
        

       
        
        # Filtrar por  campos que contengan la palabra clave en su nombre
        if name:
            queryset = queryset.filter(Q(name__icontains=name))
        

        
        queryset = queryset.order_by('id')
        return queryset
    
# ------------------------------------------------ #
# Paginación de certificados de disposiicón #
class FinalCertificateList_Pagination(LoginRequiredMixin,View):
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        
        filtered_name = request.session.get('filtered_name')
        
        
        url = reverse('residuos:final_certificate_view')
        params = {}
        
        if filtered_name:
            params['name'] = filtered_name
 
                
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)
    
# --------------------------------------- #
# Editar Certificado de Disposición Final #
class EditFinalCertificate(LoginRequiredMixin, View):
    template_name = 'UniCLab_Residuos/editar_certificado_disposicion.html'

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        try:
            # Decodificar el ID en base64 dos veces
            certificado_key = kwargs.get('pk')
            certificado_id = base64.urlsafe_b64decode(base64.urlsafe_b64decode(certificado_key)).decode('utf-8')
            # Obtener el registro de residuos que se va a editar
            certificado = get_object_or_404(CERTIFICADO_DISPOSICION, id=certificado_id)
            # Crear el formulario con los datos del registro
            form = CertificadoDisposicionForm(instance=certificado)
            today=date.today().isoformat()
            date_cert=certificado.date
            date_cert = certificado.date.strftime('%Y-%m-%d')  # Formatear la fecha correctamente
            
           
            return render(request, self.template_name, {'form': form, 'today':today, 'date_cert':date_cert,})
        except Exception as e:
            print(e)
            return HttpResponseBadRequest('Error interno del servidor')

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Decodificar el ID en base64 dos veces
            certificado_key = kwargs.get('pk')
            certificado_id = base64.urlsafe_b64decode(base64.urlsafe_b64decode(certificado_key)).decode('utf-8')
            # Obtener el registro de residuos que se va a editar
            certificado = get_object_or_404(CERTIFICADO_DISPOSICION, id=certificado_id)
            
            # Crear el formulario con los datos del registro
            form = CertificadoDisposicionForm(request.POST, request.FILES, instance=certificado)
            
            
            if form.is_valid():
                # Asignar el usuario actual a last_updated_by
                form.instance.last_updated_by = request.user
                
                certificado = form.save()

                # Registrar evento
                tipo_evento = 'EDICION CERTIFICADO DE DISPOSICION FINAL'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = 'Cambios actualizados correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                return JsonResponse({'success': False, 'errors': form.errors})
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor:{mensaje}')

# ------------------------------------- #
# Desactivar Certificado de Disposición #
class DisableFinalDispositionCertificate(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            certificado_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(certificado_key).decode('utf-8')
            certificado_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia del registro
            certificado = get_object_or_404(CERTIFICADO_DISPOSICION, id=certificado_id)
            # incluir el usuario que realiza el cambio
            certificado.last_updated_by= request.user
            # Desactivar el registro
            certificado.is_active = False
            certificado.save()
            
            # Registrar evento
            tipo_evento = 'DESHABILITAR CERTIFICADO DE DISPOSICION FINAL'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Registro de certificado deshabilitado correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})
        
# ------------------------------------- #
# Desactivar Certificado de Disposición #
class EnableFinalDispositionCertificate(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            certificado_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(certificado_key).decode('utf-8')
            certificado_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia del registro
            certificado = get_object_or_404(CERTIFICADO_DISPOSICION, id=certificado_id)
            # incluir el usuario que realiza el cambio
            certificado.last_updated_by= request.user
            # Desactivar el registro
            certificado.is_active = True
            certificado.save()
            
            # Registrar evento
            tipo_evento = 'HABILITAR CERTIFICADO DE DISPOSICION FINAL'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Registro de certificado habilitado correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})
        

class CreateImportantLinks(LoginRequiredMixin, View):
    template_name = 'UniCLab_Residuos/crear_enlace_interes.html'

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        form = InformacionInteresForm()
        return render(request, self.template_name, {'form': form})

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        form = InformacionInteresForm(request.POST, request.FILES)
        try:
            if form.is_valid():
                # Asignar el usuario actual a los campos created_by y last_updated_by
                form.instance.created_by = request.user
                form.instance.last_updated_by = request.user

                # Si el tipo es "Video de Youtube", extraer el ID del video
                if form.cleaned_data['tipo'] == 'Video de Youtube':
                    youtube_url = form.cleaned_data['url']
                    youtube_id = self.extract_youtube_id(youtube_url)
                    form.instance.id_youtube = youtube_id

                # Guardar el registro si el formulario es válido
                enlace_interes = form.save()
                
                # Registrar evento
                tipo_evento = 'CREAR MATERIAL DE INTERES'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = f'Material de interés creado correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                print(form.errors)
                return JsonResponse({'success': False, 'errors': form.errors})
                
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor: {mensaje}')

    def extract_youtube_id(self, url):
        """
        Extract the YouTube video ID from the given URL.
        """
        # Regular expression to extract the video ID from YouTube URL
        pattern = r'(?:v=|\/)([0-9A-Za-z_-]{11}).*'
        match = re.search(pattern, url)
        if match:
            return match.group(1)
        return None



# ------------------------------------ #
# Visualización de Material de Interés #
class ViewImportantLinks(LoginRequiredMixin,ListView):

    model = InformacionInteres
    template_name = "UniCLab_Residuos/listado_informacion_interes.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        name = self.request.GET.get('name')
        # Guardar los valores de filtrado en la sesión
        request.session['filtered_name'] = name
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['usuarios'] = User.objects.all()
        
       
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        name = self.request.GET.get('name')
        
        # Filtrar por  campos que contengan la palabra clave en su nombre
        if name:
            queryset = queryset.filter(Q(name__icontains=name))
        

        
        queryset = queryset.order_by('-id')
        return queryset


# ------------------------------------------------ #
# Paginación de material de interés #
class ImportantLinks_Pagination(LoginRequiredMixin,View):
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        
        filtered_name = request.session.get('filtered_name')
        
        
        url = reverse('residuos:important_link_view')
        params = {}
        
        if filtered_name:
            params['name'] = filtered_name
 
                
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)
    
# --------------------------------------- #
# Editar Certificado de Disposición Final #
class EditImportantLink(LoginRequiredMixin, View):
    template_name = 'UniCLab_Residuos/editar_enlace_interes.html'

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, *args, **kwargs):
        try:
            # Decodificar el ID en base64 dos veces
            link_key = kwargs.get('pk')
            link_id = base64.urlsafe_b64decode(base64.urlsafe_b64decode(link_key)).decode('utf-8')
            # Obtener el registro de residuos que se va a editar
            link = get_object_or_404(InformacionInteres, id=link_id)
            # Crear el formulario con los datos del registro
            form = InformacionInteresForm(instance=link)
            
            
           
            return render(request, self.template_name, {'form': form, })
        except Exception as e:
            print(e)
            return HttpResponseBadRequest('Error interno del servidor')

    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Decodificar el ID en base64 dos veces
            link_key = kwargs.get('pk')
            link_id = base64.urlsafe_b64decode(base64.urlsafe_b64decode(link_key)).decode('utf-8')
            # Obtener el registro de residuos que se va a editar
            link = get_object_or_404(InformacionInteres, id=link_id)
            
            # Crear el formulario con los datos del registro
            form = InformacionInteresForm(request.POST, request.FILES, instance=link)
            
            
            if form.is_valid():
                # Asignar el usuario actual a last_updated_by
                form.instance.last_updated_by = request.user
                
                link = form.save()

                # Registrar evento
                tipo_evento = 'EDICION MATERIAL DE INTERES'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = 'Cambios actualizados correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                return JsonResponse({'success': False, 'errors': form.errors})
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor:{mensaje}')

# ------------------------------- #
# Deshabilitar Enlaces de interés #
class DisableImportantLink(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            link_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(link_key).decode('utf-8')
            link_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia del registro
            link = get_object_or_404(InformacionInteres, id=link_id)
            # incluir el usuario que realiza el cambio
            link.last_updated_by= request.user
            # Desactivar el registro
            link.is_active = False
            link.save()
            
            # Registrar evento
            tipo_evento = 'DESHABILITAR MATERIAL DE INTERES'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Registro de material de interés deshabilitado correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})
        
# ------------------------------- #
# Habilitar Enlaces de interés #
class EnableImportantLink(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR', 'ADMINISTRADOR AMBIENTAL'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            link_key = kwargs.get('pk')
            item_id_encoded = base64.urlsafe_b64decode(link_key).decode('utf-8')
            link_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia del registro
            link = get_object_or_404(InformacionInteres, id=link_id)
            # incluir el usuario que realiza el cambio
            link.last_updated_by= request.user
            # Desactivar el registro
            link.is_active = True
            link.save()
            
            # Registrar evento
            tipo_evento = 'HABILITAR MATERIAL DE INTERES'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Registro de material de interés habilitado correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})


# ------------------------------------------------------------------ #
# Exportar a Excel registro de residuos por solicitud en formato RP1 #
class Export2ExcelWastesRP1(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
       
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO', 'ADMINISTRADOR AMBIENTAL'])
    def get(self, request, solicitud_id, *args, **kwargs):

        def get_file_name():
            now = datetime.now()
            return f'Mz.FT.SGA.006 Formato RP1 Información Residuos Almacenamiento Temporal {now.strftime("%d%m%Y%H%M%S")}.xlsx'
        
        
        item_id_encoded = base64.urlsafe_b64decode(solicitud_id).decode('utf-8')
        solicitud = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
        registro_solicitud=get_object_or_404(SOLICITUD_RESIDUO, id=solicitud)
        

        queryset=REGISTRO_RESIDUOS.objects.filter(registro_solicitud=solicitud, residuo_enviado=True)        
        
        queryset = queryset.order_by('id')

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Ajustar la altura de la fila 1
        sheet.row_dimensions[1].height = 90

        # Ajustar la altura de la fila 3
        sheet.row_dimensions[3].height = 20
        # Ajustar la altura de la fila 5
        sheet.row_dimensions[5].height = 20
        # Ajustar la altura de la fila 7
        sheet.row_dimensions[7].height = 20

        # Unir celdas para el texto de la primera línea
        sheet.merge_cells('A1:D1')
        cell = sheet['A1']
        cell.value = 'Sistema de Gestión Ambiental\nFormato RP1 Información Residuos Almacenamiento Temporal'
        cell.font = Font(bold=True, italic=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

        # Unir celdas para la imagen
        # sheet.merge_cells('E1:H1')
        logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')
        pil_image = PILImage.open(logo_path)
        image = ExcelImage(pil_image)
        
        
        # Agregar la imagen a la celda unida
        sheet.add_image(image, 'F1')

     

        # Unir celdas para el texto de la segunda línea
        sheet.merge_cells('A2:H2')
        cell = sheet['A2']
        cell.value = 'Dependencia / Area / Laboratorio:'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
        

        # Unir celdas para el texto de la tercera  línea
        sheet.merge_cells('A3:H3')
        cell = sheet['A3']
        cell.value = f'{registro_solicitud.created_by.lab.name}'
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='bottom')

        # Bordes para líneas 2 y 3
        sheet['A2'].border=Border(left=Side(style='thin'),top=Side(style='thin'),)
        sheet['B2'].border=Border(top=Side(style='thin'),)
        sheet['C2'].border=Border(top=Side(style='thin'),)
        sheet['D2'].border=Border(top=Side(style='thin'),)
        sheet['E2'].border=Border(top=Side(style='thin'),)
        sheet['F2'].border=Border(top=Side(style='thin'),)
        sheet['G2'].border=Border(top=Side(style='thin'),)
        sheet['H2'].border=Border(top=Side(style='thin'),right=Side(style='thin'),)

        sheet['A3'].border=Border(left=Side(style='thin'),bottom=Side(style='thin'),)
        sheet['B3'].border=Border(bottom=Side(style='thin'),)
        sheet['C3'].border=Border(bottom=Side(style='thin'),)
        sheet['D3'].border=Border(bottom=Side(style='thin'),)
        sheet['E3'].border=Border(bottom=Side(style='thin'),)
        sheet['F3'].border=Border(bottom=Side(style='thin'),)
        sheet['G3'].border=Border(bottom=Side(style='thin'),)
        sheet['H3'].border=Border(bottom=Side(style='thin'),right=Side(style='thin'),)

        # Unir celdas para el texto de la cuarta línea
        sheet.merge_cells('A4:D4')
        cell = sheet['A4']
        cell.value = 'Ubicación de la Dependencia / Area / Laboratorio: '
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
        sheet.merge_cells('E4:H4')
        cell = sheet['E4']
        cell.value = '(Bloque, piso, oficina, campus, otros)'
        cell.font = Font(bold=False, italic=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

        # Unir celdas para el texto de la quinta  línea
        sheet.merge_cells('A5:H5')
        cell = sheet['A5']
        cell.value = f'{registro_solicitud.created_by.lab.campus_location}'
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='bottom')

        # Bordes para líneas 4 y 5
        sheet['A4'].border=Border(left=Side(style='thin'),)
        sheet['H4'].border=Border(right=Side(style='thin'),)
        sheet['A5'].border=Border(left=Side(style='thin'),bottom=Side(style='thin'),)
        sheet['B5'].border=Border(bottom=Side(style='thin'),)
        sheet['C5'].border=Border(bottom=Side(style='thin'),)
        sheet['D5'].border=Border(bottom=Side(style='thin'),)
        sheet['E5'].border=Border(bottom=Side(style='thin'),)
        sheet['F5'].border=Border(bottom=Side(style='thin'),)
        sheet['G5'].border=Border(bottom=Side(style='thin'),)
        sheet['H5'].border=Border(bottom=Side(style='thin'),right=Side(style='thin'),)

        # Unir celdas para el texto de la sexta línea
        sheet.merge_cells('A6:H6')
        cell = sheet['A6']
        cell.value = 'Nombre y cargo del responsable que entrega el residuo: '
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

        # Unir celdas para el texto de la SÉPTIMA  línea
        sheet.merge_cells('A7:H7')
        cell = sheet['A7']
        cell.value = f'{registro_solicitud.created_by.first_name} {registro_solicitud.created_by.first_name} - {registro_solicitud.created_by.position_company}'
        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='bottom')

        # Bordes para líneas 6 y 7
        sheet['A6'].border=Border(left=Side(style='thin'),)
        sheet['H6'].border=Border(right=Side(style='thin'),)
        sheet['A7'].border=Border(left=Side(style='thin'),bottom=Side(style='thin'),)
        sheet['B7'].border=Border(bottom=Side(style='thin'),)
        sheet['C7'].border=Border(bottom=Side(style='thin'),)
        sheet['D7'].border=Border(bottom=Side(style='thin'),)
        sheet['E7'].border=Border(bottom=Side(style='thin'),)
        sheet['F7'].border=Border(bottom=Side(style='thin'),)
        sheet['G7'].border=Border(bottom=Side(style='thin'),)
        sheet['H7'].border=Border(bottom=Side(style='thin'),right=Side(style='thin'),)

        # Unir celdas para el texto de la Octava  línea
        sheet.merge_cells('A8:H8')
        cell = sheet['A8']
        cell.font = Font(bold=False, italic=True)
        cell.value = f'** Decreto 1076 de 2015'
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        
        sheet.merge_cells('A9:A10')
        sheet['A9'] = 'No.'
        sheet.merge_cells('B9:B10')
        sheet['B9'] = 'Fecha diligenciamiento'
        sheet.merge_cells('C9:C10')      
        sheet['C9'] = 'Residuo'
        sheet.merge_cells('D9:D10')
        sheet['D9'] = 'Cant.'
        sheet['E9'] = 'Und.'
        sheet['E10'] = '(mg/ml)'
        sheet['F9'] = 'Clasificado:'
        sheet['F10'] = 'Y - A**'
        sheet['G9'] = 'Estado'
        sheet['G10'] = 'Líquido/Sólido'
        sheet['H9'] = 'Fecha de Salida'
        sheet['H10'] = 'Centro de Acopio'

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True)
        alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
        row = 9
        for col in range(1, 9):
            # sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).font = bold_font
            sheet.cell(row=row, column=col).alignment = alignment

        normal_italic_font = Font(bold=False, italic=True)
        alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
        row = 10
        for col in range(1, 9):
            # sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).font = normal_italic_font
            sheet.cell(row=row, column=col).alignment = alignment

        # Bordes para líneas 9 y 10
        sheet['A9'].border=Border(left=Side(style='thin'),top=Side(style='thin'),right=Side(style='thin'),)
        sheet['B9'].border=Border(top=Side(style='thin'),right=Side(style='thin'),)
        sheet['C9'].border=Border(top=Side(style='thin'),right=Side(style='thin'),)
        sheet['D9'].border=Border(top=Side(style='thin'),right=Side(style='thin'),)
        sheet['E9'].border=Border(top=Side(style='thin'),right=Side(style='thin'),)
        sheet['F9'].border=Border(top=Side(style='thin'),right=Side(style='thin'),)
        sheet['G9'].border=Border(top=Side(style='thin'),right=Side(style='thin'),)
        sheet['H9'].border=Border(top=Side(style='thin'),right=Side(style='thin'),)

        sheet['A10'].border=Border(left=Side(style='thin'),bottom=Side(style='thin'),right=Side(style='thin'),)
        sheet['B10'].border=Border(bottom=Side(style='thin'),right=Side(style='thin'),)
        sheet['C10'].border=Border(bottom=Side(style='thin'),right=Side(style='thin'),)
        sheet['D10'].border=Border(bottom=Side(style='thin'),right=Side(style='thin'),)
        sheet['E10'].border=Border(bottom=Side(style='thin'),right=Side(style='thin'),)
        sheet['F10'].border=Border(bottom=Side(style='thin'),right=Side(style='thin'),)
        sheet['G10'].border=Border(bottom=Side(style='thin'),right=Side(style='thin'),)
        sheet['H10'].border=Border(bottom=Side(style='thin'),right=Side(style='thin'),)

       

        # # Establecer Anchos personalizados
        sheet.column_dimensions['A'].width = 4
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 26
        sheet.column_dimensions['D'].width = 6
        sheet.column_dimensions['E'].width = 8
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 10


        

        row = 11
        
        for item in queryset:
            # Organizar las clasificaciones
            clasificaciones = ', '.join(item.clasificado.values_list('name', flat=True)) if item.clasificado.exists() else 'NO DEFINIDA'

            
            
            sheet.cell(row=row, column=1).value = f'{row-10}'
            sheet.cell(row=row, column=1).font = Font(bold=True)
            sheet.cell(row=row, column=2).value = item.date_create.strftime("%d/%m/%Y")
            sheet.cell(row=row, column=3).value = item.nombre_residuo
            sheet.cell(row=row, column=4).value = item.total_residuo
            sheet.cell(row=row, column=5).value = item.unidades.name
            sheet.cell(row=row, column=6).value = clasificaciones
            sheet.cell(row=row, column=7).value = item.estado.name
            sheet.cell(row=row, column=8).value = ''
            
            for col in range(1, 9):
                sheet.cell(row=row, column=col).border = thin_border
                sheet.cell(row=row, column=col).alignment = alignment

            row += 1
        # Después de la última fila de la tabla, en la columna A
        sheet.cell(row=row, column=1).value = "Código: Mz.FT.SGA.006"
        sheet.cell(row=row, column=1).font = Font(bold=True, italic=True)
        sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left')

        # En la columna H
        sheet.cell(row=row, column=8).value = "Version: 1.0"
        sheet.cell(row=row, column=8).font = Font(bold=True, italic=True)
        sheet.cell(row=row, column=8).alignment = Alignment(horizontal='right')
        sheet.row_dimensions[row].height = 22  # Establecer el alto de la fila en 22

        start_column = 1
        end_column = 8
        start_row = 4
        end_row = row - 1

        start_column_letter = get_column_letter(start_column)
        end_column_letter = get_column_letter(end_column)

        table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

        

        fill = PatternFill(fill_type="solid", fgColor=WHITE)
        start_cell = sheet['A1']
        end_column_letter = get_column_letter(end_column)
        end_row = row+1
        end_cell = sheet[end_column_letter + str(end_row)]
        table_range = start_cell.coordinate + ':' + end_cell.coordinate

        for row in sheet[table_range]:
            for cell in row:
                cell.fill = fill

        # Color de fondo deseado en formato hexadecimal
        color_hex = "d9d9d9"

        # Crear un objeto PatternFill con el color deseado
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        # Aplicar el color de fondo al rango de celdas A9:H10
        for row in range(9, 11):
            for column in range(1, 9):
                sheet.cell(row=row, column=column).fill = fill

        # Color de texto deseado en formato hexadecimal
        color_hex = "FF0000"  # Rojo

        # Crear un objeto Font con el color deseado
        font = Font(color=color_hex, bold=True)

        # Aplicar el color de texto a la celda B9
        sheet['B9'].font = font

        

        # Establecer el tamaño de la hoja para impresión (Letter)
        sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER

        # Configuración de los márgenes
        workbook.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0.1, footer=0.1)

        # Establecer el ancho total de la hoja para que se ajuste a las márgenes
        sheet.page_setup.fitToWidth = 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={get_file_name()}'

        workbook.save(response)

        tipo_evento = 'DESCARGA DE ARCHIVOS'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        return response