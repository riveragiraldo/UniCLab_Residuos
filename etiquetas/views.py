from django.shortcuts import render

from reactivos.views import *
from residuos.views import *
from reactivos.models import *
from residuos.models import *
from .models import *
from .forms import*
from .utils import check_group_permission


from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintPageSetup
from PIL import Image as PILImage
import openpyxl
from openpyxl import Workbook
import math
from openpyxl.worksheet.page import PageMargins

# -------------------------------------------- #
# Vista para la descarga del formato en blanco #
class DownloadFormatLabelView(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def get(self, request, *args, **kwargs):
        try:
            configuracion_sistema = ConfiguracionSistema.objects.first()

            if configuracion_sistema and configuracion_sistema.formato_libre:
                formato = configuracion_sistema.formato_libre
                filename = formato.name.split('/')[-1]

                # Crear una respuesta de archivo directamente con FileResponse
                response = FileResponse(formato.open(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'

                # Registro de mensaje (opcional)
                mensaje = f'Se ha descargado el archivo {filename}'
                print(mensaje)

                return response
            else:
                return HttpResponse("El formato no está disponible en este momento.", status=404)
        except Exception as e:
            return HttpResponse(f"Error al descargar el formato: {str(e)}", status=500)


# ----------------------------------------------------------------- #
# Vista para la generación de las etiquetas de susrtancias químicas #

class GenerateLabel(LoginRequiredMixin, View):
    template_name = 'Etiquetas/generar_etiquetas.html'
    
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def get(self, request, *args, **kwargs):
        context = {
            'labs': Laboratorios.objects.all(),
        }
        return render(request, self.template_name, context)

    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def post(self, request, *args, **kwargs):
        
        try:
            # Obtener los valores de los campos del formulario
            nombre = request.POST.get('reagent')
            numero = request.POST.get('number')
            cas = request.POST.get('cas')
            id_sustancia = request.POST.get('id_reagent')
            etiqueta = request.POST.get('label_type')
            laboratorio = request.POST.get('lab')
            id_hermes = request.POST.get('ID_Hermes')
            telefono_emergencia = request.POST.get('emergency_number')
            # Verificar estado del checkbox
            formato_libre = request.POST.get('empty_template')
            formato_libre_estado = True if formato_libre == 'on' else False


            if not id_hermes:
                id_hermes=''
            if not laboratorio:
                laboratorio=''
            if not telefono_emergencia:
                telefono_emergencia=''

            # Imprimir los valores en la terminal
            print("Nombre:", nombre)
            print("Número: ", numero)
            print("CAS:", cas)
            print("ID Sustancia:", id_sustancia)
            print("Etiqueta:", etiqueta)
            print("Laboratorio:", laboratorio)
            print("ID Hermes:", id_hermes)
            print("Teléfono Emergencia:", telefono_emergencia)
            print("Formato Libre:", formato_libre_estado)
            # try:
            # Si formato_libre_estado es True, descargar el formato libre
            if formato_libre_estado:
                mensaje='Generando'

                print(mensaje)
                # return redirect('etiquetas:download_format')
                return JsonResponse({'success': True, 'message': mensaje})

            # Verificar campos necesarios
            if not nombre or not id_sustancia or not etiqueta:
                mensaje = 'Error: Los campos Nombre, ID Sustancia y Etiqueta son obligatorios.'
                print(mensaje)
                return JsonResponse({'success': False, 'errors': mensaje})

            # Verificación adicional para el número de teléfono
            if telefono_emergencia and not re.match(r'^(60\d{8}|3\d{9})$', telefono_emergencia):
                mensaje = 'Error: Ingrese un número telefónico válido para Colombia (60xxxxxxxx o 3xxxxxxxxx).'
                print(mensaje)
                return JsonResponse({'success': False, 'message': mensaje})

            # Funcionalidad basada en la etiqueta seleccionada
            if etiqueta == '1':
                numero=int(numero)
                if numero > 56:
                    mensaje= f'Solo es posible generar 56 etiquetas básicas de manera simultanea, revise e intente nuevamente'
                    return JsonResponse({'success': False, 'errors': mensaje})

                mensaje = 'Enlace de etiqueta básica generado correctamente'
                print(mensaje)
                download_url = reverse('etiquetas:download_basic_label')
                # Agregar parámetros a la URL
                download_url = f"{download_url}?substance_name={nombre}&substance_id={id_sustancia}&label_number={numero}"
                print(download_url)
                messages.success(self.request, f'{download_url}')
                return JsonResponse({'success': True, 'message': mensaje})



            elif etiqueta == '2':
                numero=int(numero)
                if numero > 8:
                    mensaje= f'Solo es posible generar 8 etiquetas medianas de manera simultanea, revise e intente nuevamente'
                    return JsonResponse({'success': False, 'errors': mensaje})

                mensaje = 'Enlace de etiqueta mediana generado correctamente'
                print(mensaje)
                download_url = reverse('etiquetas:download_medium_label')
                # Agregar parámetros a la URL
                download_url = f"{download_url}?substance_name={nombre}&substance_id={id_sustancia}&label_number={numero}&lab={laboratorio}&hermes_id={id_hermes}&e_number={telefono_emergencia}"
                print(download_url)
                messages.success(self.request, f'{download_url}')
                return JsonResponse({'success': True, 'message': mensaje})
            elif etiqueta == '3':
                numero=int(numero)
                if numero > 24:
                    mensaje= f'Solo es posible generar 24 etiquetas pequeñas de manera simultanea, revise e intente nuevamente'
                    return JsonResponse({'success': False, 'errors': mensaje})

                mensaje = 'Enlace de etiqueta pequeña generado correctamente'
                print(mensaje)
                download_url = reverse('etiquetas:download_small_label')
                # Agregar parámetros a la URL
                download_url = f"{download_url}?substance_name={nombre}&substance_id={id_sustancia}&label_number={numero}&lab={laboratorio}&hermes_id={id_hermes}&e_number={telefono_emergencia}"
                print(download_url)
                messages.success(self.request, f'{download_url}')
                return JsonResponse({'success': True, 'message': mensaje})
            elif etiqueta == '4':
                numero=int(numero)
                if numero > 4:
                    mensaje= f'Solo es posible generar 4 etiquetas grandes de manera simultanea, revise e intente nuevamente'
                    return JsonResponse({'success': False, 'errors': mensaje})

                mensaje = 'Enlace de etiqueta grande generado correctamente'
                print(mensaje)
                download_url = reverse('etiquetas:download_big_label')
                # Agregar parámetros a la URL
                download_url = f"{download_url}?substance_name={nombre}&substance_id={id_sustancia}&label_number={numero}&lab={laboratorio}&hermes_id={id_hermes}&e_number={telefono_emergencia}"
                print(download_url)
                messages.success(self.request, f'{download_url}')
                return JsonResponse({'success': True, 'message': mensaje})

            # Si no se cumple ninguno de los casos de etiqueta
            mensaje = 'Error: Tipo de etiqueta no válido.'
            print(mensaje)
            return JsonResponse({'success': False, 'message': mensaje})
        
        except Exception as e:
            # Manejo de cualquier excepción no prevista
            mensaje = f'Error: Ocurrió un problema inesperado. Detalles: {str(e)}'
            print(mensaje)
            return JsonResponse({'success': False, 'message': mensaje})

                
                
           
            
           
    
# ------------------------------------ #
# Autocompletado de sustancias químicas #

class autocompleteChemicalSubstances(LoginRequiredMixin,View):
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def get(self, request):
        term = request.GET.get('term', '')
        
        substances = Sustancias.objects.filter(
                Q(name__icontains=term) | Q(cas__icontains=term)).order_by('name')[:20]        

        results = []
        for substance in substances:
            result = {
                'name': substance.name,
                'cas': substance.cas,
                'id_substance':substance.id,
            }
            results.append(result)
        return JsonResponse(results, safe=False)



# ----------------------- #
# Generar Etiqueta Grande #
class GenerateBigLabel(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
       
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def get(self, request, *args, **kwargs):
        def get_file_name():
            nombre_sustancia = request.GET.get('substance_name')
            now = datetime.now()
            return f'Etiqueta_grande_{nombre_sustancia}_{now.strftime("%d%m%Y%H%M%S")}.xlsx'
        
        # Extraer parámetros de la URL
        nombre_sustancia = request.GET.get('substance_name')
        id_sustancia = request.GET.get('substance_id')
        
        numero_etiquetas = int(request.GET.get('label_number'))  # Convertir a entero
        id_hermes = request.GET.get('hermes_id')
        lab = request.GET.get('lab')
        if lab =='':
            lab=''
        else:
            lab=get_object_or_404(Laboratorios, id=lab)
        numero_emergencia = request.GET.get('e_number')        

        # Verificar si el número de etiquetas es superior a 56
        if numero_etiquetas > 4:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Mensaje de error
            error_message = "¡Upps algo ocurrió mal!\nEl número máximo de etiquetas debe ser menor o igual a 4, revisa y genera nuevamente tus etiquetas."

            # Ajustar la celda para el mensaje
            cell = sheet.cell(row=1, column=1)
            cell.value = error_message
            cell.font = Font(name='Arial', size=10, bold=True, color='FF0000')  # Letras llamativas, tamaño más pequeño
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Ajuste de texto y nueva línea
            sheet.merge_cells('A1:E10')  # Unir celdas para el mensaje, aumentando la altura
            sheet.row_dimensions[1].height = 40  # Ajustar la altura de la fila
            sheet.column_dimensions[get_column_letter(1)].width = 50  # Ajustar el ancho de la columna
            
            # Guardar el archivo en la memoria y preparar la respuesta HTTP
            file_name = get_file_name()
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={file_name}'
            workbook.save(response)
            return response

        sustancia = get_object_or_404(Sustancias, id=id_sustancia)
        cas = sustancia.cas
        if sustancia.warning:
            advertencia = sustancia.warning.name
        else:
            advertencia=''
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Configurar la orientación de la página
        # sheet.page_setup = PrintPageSetup(orientation='landscape')

        # Definir un relleno con fondo blanco
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # Crear un borde de puntos
        thin_border = Side(style='thin')

        # Calcular número de líneas necesarias
        num_lines = math.ceil(numero_etiquetas / 1)
        

        for line in range(num_lines):
            # Calcular el offset de fila para cada etiqueta
            row_offset = line * 19  # 19 filas por etiqueta (18 etiqueta + 1 separación)

            for i in range(1): # Número de etiquetas horizontales
                index = line * 1 + i
                if index >= numero_etiquetas:
                    break

                # Calcular el offset de columna para cada etiqueta
                col_offset = i * 18
                
                # Aplicar el fondo blanco a las celdas de la etiqueta actual
                # for row in range(1, 19):  # Filas 1 a 19
                #     for col in range(1, 18):  # Columnas A a R
                #         cell = sheet.cell(row=row + row_offset, column=col + col_offset)
                #         cell.fill = white_fill

                # Anchos de columna para cada etiqueta
                sheet.column_dimensions[get_column_letter(1  + col_offset)].width = 1.90    # Ancho A = 1.27
                sheet.column_dimensions[get_column_letter(2  + col_offset)].width = 4.00    # Ancho B = 3.36
                sheet.column_dimensions[get_column_letter(3  + col_offset)].width = 7.55    # Ancho C = 6.91 
                sheet.column_dimensions[get_column_letter(4  + col_offset)].width = 7.55    # Ancho D = 6.91
                sheet.column_dimensions[get_column_letter(5  + col_offset)].width = 1.00    # Ancho E = 0.61 
                sheet.column_dimensions[get_column_letter(6  + col_offset)].width = 7.55    # Ancho F = 6.91
                sheet.column_dimensions[get_column_letter(7  + col_offset)].width = 7.55    # Ancho G = 6.91
                sheet.column_dimensions[get_column_letter(8  + col_offset)].width = 1.00    # Ancho H = 0.61
                sheet.column_dimensions[get_column_letter(9  + col_offset)].width = 7.55    # Ancho I = 6.91
                sheet.column_dimensions[get_column_letter(10 + col_offset)].width = 7.55    # Ancho J = 6.91
                sheet.column_dimensions[get_column_letter(11 + col_offset)].width = 1.00    # Ancho K = 0.61
                sheet.column_dimensions[get_column_letter(12 + col_offset)].width = 7.55    # Ancho L = 6.91
                sheet.column_dimensions[get_column_letter(13 + col_offset)].width = 7.55    # Ancho M = 6.91
                sheet.column_dimensions[get_column_letter(14 + col_offset)].width = 1.00    # Ancho N = 0.61
                sheet.column_dimensions[get_column_letter(15 + col_offset)].width = 7.55    # Ancho O = 6.91
                sheet.column_dimensions[get_column_letter(16 + col_offset)].width = 7.55    # Ancho P = 6.91
                sheet.column_dimensions[get_column_letter(17 + col_offset)].width = 4.00    # Ancho Q = 3.36
                sheet.column_dimensions[get_column_letter(18 + col_offset)].width = 1.90    # Ancho R = 1.27
                
                
                
                
                # Ajustar las alturas de las filas
                sheet.row_dimensions[1  + row_offset].height = 11.30    # Alto 1 = 11.3
                sheet.row_dimensions[2  + row_offset].height = 51.00    # Alto 2 = 51
                sheet.row_dimensions[3  + row_offset].height = 28.00    # Alto 3 = 28   
                sheet.row_dimensions[4  + row_offset].height = 23.00    # Alto 4 = 23 
                sheet.row_dimensions[5  + row_offset].height = 23.00    # Alto 5 = 23
                sheet.row_dimensions[6  + row_offset].height = 5.800    # Alto 6 = 5.8  
                sheet.row_dimensions[7  + row_offset].height = 85.00    # Alto 7 = 85  
                sheet.row_dimensions[8  + row_offset].height = 5.800    # Alto 8 = 5.8 
                sheet.row_dimensions[9  + row_offset].height = 23.00    # Alto 9 = 23
                sheet.row_dimensions[10 + row_offset].height = 39.00    # Alto 10 = 39
                sheet.row_dimensions[11 + row_offset].height = 20.00    # Alto 11 = 20
                # sheet.row_dimensions[12 + row_offset].height = 239.0    # Alto 12 = 239
                sheet.row_dimensions[13 + row_offset].height = 21.00    # Alto 13 = 21
                sheet.row_dimensions[14 + row_offset].height = 21.00    # Alto 14 = 21
                sheet.row_dimensions[15 + row_offset].height = 21.00    # Alto 15 = 21
                sheet.row_dimensions[16 + row_offset].height = 21.00    # Alto 16 = 21
                sheet.row_dimensions[17 + row_offset].height = 21.00    # Alto 17 = 21
                sheet.row_dimensions[18 + row_offset].height = 11.30    # Alto 18 = 11.3
                

                # Separador para la siguiente etiqueta abajo
                sheet.row_dimensions[19+ row_offset].height = 5.8       # Alto 19 = 5.8           
                
                # Aplicar bordes manualmente para formar el rectángulo
                # A1: Borde izquierdo y borde arriba
                sheet.cell(row=1 + row_offset, column=1 + col_offset).border = Border(left=thin_border, top=thin_border)
                # A2:A17 Borde izquierdo
                for row in range(2, 18):
                    sheet.cell(row=row + row_offset, column=1 + col_offset).border = Border(left=thin_border)

                # A18 Borde izquierdo y borde abajo
                sheet.cell(row=18 + row_offset, column=1 + col_offset).border = Border(left=thin_border, bottom=thin_border)

                # B1:Q1 Borde arriba
                for col in range(2, 18):
                    sheet.cell(row=1 + row_offset, column=col + col_offset).border = Border(top=thin_border)

                # R1: Borde arriba, borde derecho
                sheet.cell(row=1 + row_offset, column=18 + col_offset).border = Border(top=thin_border, right=thin_border)

                # B18:Q18 Borde abajo
                for col in range(2, 18):
                    sheet.cell(row=18 + row_offset, column=col + col_offset).border = Border(bottom=thin_border)

                # Rl8 Borde abajo, borde derecho
                sheet.cell(row=18 + row_offset, column=18 + col_offset).border = Border(bottom=thin_border, right=thin_border)

                # R2:R17 Borde derecho
                for row in range(2, 18):
                    sheet.cell(row=row + row_offset, column=18 + col_offset).border = Border(right=thin_border)

                #### Unir celdas B2:Q2 y colocar "Nombre de la sustancia" ####
                sheet.merge_cells(start_row=2 + row_offset, start_column=2 + col_offset, end_row=2 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=2 + row_offset, column=2 + col_offset)
                merged_cell.value = nombre_sustancia.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=34, bold=True)  # Fuente, tamaño y negrita

                
                # Unir celdas B3:I3 y colocar "CAS"
                sheet.merge_cells(start_row=3 + row_offset, start_column=2 + col_offset, end_row=3 + row_offset, end_column=9 + col_offset)
                merged_cell = sheet.cell(row=3 + row_offset, column=2 + col_offset)
                merged_cell.value = 'CAS'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=18, bold=True)  # Fuente y tamaño

                # Unir celdas J3:Q3 y colocar CAS de la sustancia
                sheet.merge_cells(start_row=3 + row_offset, start_column=10 + col_offset, end_row=3 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=3 + row_offset, column=10 + col_offset)
                merged_cell.value = cas.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=18, bold=True)  # Fuente y tamaño

                # Unir celdas B4:Q4 y colocar "Etiquetado conforme al SGA/GHS:"
                sheet.merge_cells(start_row=4 + row_offset, start_column=2 + col_offset, end_row=4 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=4 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Etiquetado conforme al SGA/GHS:'  
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=14, underline='single', bold=True)  # Fuente, tamaño y negrita
    

                # Unir celdas B5:Q5 y colocar "Pictogramas"
                sheet.merge_cells(start_row=5 + row_offset, start_column=2 + col_offset, end_row=5 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=5 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Pictográmas'  
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=14, underline='single', bold=True)  # Fuente, tamaño y negrita

                # C7: Agregar el pictograma1 si existe
                if sustancia and sustancia.pictogram_clp1:
                    fila_picto=7+row_offset
                    pictogram_path = sustancia.pictogram_clp1.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 113.2     # Ajuste Altura 3.0
                    img.width = 113.2      # Ajuste Ancho 3.0
                    img.anchor = f'{get_column_letter(3 + col_offset)}{fila_picto}'
                    sheet.add_image(img)

                # F7: Agregar el pictograma2 si existe
                if sustancia and sustancia.pictogram_clp2:
                    fila_picto=7+row_offset
                    pictogram_path = sustancia.pictogram_clp2.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 113.2     # Ajuste Altura 3.0
                    img.width = 113.2      # Ajuste Ancho 3.0
                    img.anchor = f'{get_column_letter(6 + col_offset)}{fila_picto}'
                    sheet.add_image(img)
                
                # I7: Agregar el pictograma3 si existe
                if sustancia and sustancia.pictogram_clp3:
                    fila_picto=7+row_offset
                    pictogram_path = sustancia.pictogram_clp3.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 113.2     # Ajuste Altura 3.0
                    img.width = 113.2      # Ajuste Ancho 3.0
                    img.anchor = f'{get_column_letter(9 + col_offset)}{fila_picto}'
                    sheet.add_image(img)
                
                # L7: Agregar el pictograma4 si existe
                if sustancia and sustancia.pictogram_clp4:
                    fila_picto=7+row_offset
                    pictogram_path = sustancia.pictogram_clp4.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 113.2     # Ajuste Altura 3.0
                    img.width = 113.2      # Ajuste Ancho 3.0
                    img.anchor = f'{get_column_letter(12 + col_offset)}{fila_picto}'
                    sheet.add_image(img)
                
                # O7: Agregar el pictograma5 si existe
                if sustancia and sustancia.pictogram_clp5:
                    fila_picto=7+row_offset
                    pictogram_path = sustancia.pictogram_clp5.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 113.2     # Ajuste Altura 3.0
                    img.width = 113.2      # Ajuste Ancho 3.0
                    img.anchor = f'{get_column_letter(15 + col_offset)}{fila_picto}'
                    sheet.add_image(img)

                # Unir celdas B9:Q9 y colocar "Palabra de Advertencia"
                sheet.merge_cells(start_row=9 + row_offset, start_column=2 + col_offset, end_row=9 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=9 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Palabra de Advertencia:'  
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=14, bold=True)  # Fuente, tamaño y negrita

                # Unir celdas B10:Q10 y colocar "La Palabra de Advertencia"
                sheet.merge_cells(start_row=10 + row_offset, start_column=2 + col_offset, end_row=10 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=10 + row_offset, column=2 + col_offset)
                merged_cell.value = advertencia.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=28, bold=True, color="FF0000")  # Fuente, tamaño y negrita
                
                
                # Unir celdas B11:I11 y Colocar Frases H
                sheet.merge_cells(start_row=11 + row_offset, start_column=2 + col_offset, end_row=11 + row_offset, end_column=9 + col_offset)
                merged_cell = sheet.cell(row=11 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Frases H'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=15, bold=True)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(2, 10):
                    cell = sheet.cell(row=11 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)
                
                # Unir celdas J11:Q11 y Colocar Frases P
                sheet.merge_cells(start_row=11 + row_offset, start_column=10 + col_offset, end_row=11 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=11 + row_offset, column=10 + col_offset)
                merged_cell.value = 'Frases P'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=15, bold=True)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(10, 18):
                    cell = sheet.cell(row=11 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)

                # Organizar Frases H
                frases_h = '\n'.join(sustancia.phrase_h.values_list('name', flat=True)) if sustancia.phrase_h.exists() else '-'

                # Unir celdas B12:I12 y Colocar las Frases H
                sheet.merge_cells(start_row=12 + row_offset, start_column=2 + col_offset, end_row=12 + row_offset, end_column=9 + col_offset)
                merged_cell = sheet.cell(row=12 + row_offset, column=2 + col_offset)
                merged_cell.value = frases_h
                merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño

                # Configurar bordes superior e inferior
                for col in range(2, 10):
                    cell = sheet.cell(row=12 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)
                    

                # Organizar Frases P
                frases_p = '\n'.join(sustancia.phrase_p.values_list('name', flat=True)) if sustancia.phrase_p.exists() else '-'

                # Unir celdas J12:Q12 y Colocar las Frases P
                sheet.merge_cells(start_row=12 + row_offset, start_column=10 + col_offset, end_row=12 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=12 + row_offset, column=10 + col_offset)
                merged_cell.value = frases_p
                merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)  # Ajuste de texto
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño

                # Configurar bordes superior e inferior
                for col in range(10, 18):
                    cell = sheet.cell(row=12 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)
                
                # Ajustar altura de las filas según el contenido
                if not frases_h=='-' or not frases_p=='-':
                    max_lines_h = max(len(frases_h.split('\n')), 1)  # Calcular el número de líneas para frases H
                    max_lines_p = max(len(frases_p.split('\n')), 1)  # Calcular el número de líneas para frases P
                    max_lines = max(max_lines_h, max_lines_p)  # Obtener el mayor número de líneas entre H y P
                    line_height = 17  # Altura estimada por línea 
                    total_height = max_lines * line_height * 2

                    # Ajustar las alturas de la fila 12
                    sheet.row_dimensions[12 + row_offset].height = total_height

                # Unir celdas B13:C13 y Colocar Cantidad
                sheet.merge_cells(start_row=13 + row_offset, start_column=2 + col_offset, end_row=13 + row_offset, end_column=3 + col_offset)
                merged_cell = sheet.cell(row=13 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Cantidad:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño

                # Unir celdas D13:F13 y Colocar la raya
                sheet.merge_cells(start_row=13 + row_offset, start_column=4 + col_offset, end_row=13 + row_offset, end_column=6 + col_offset)
                merged_cell = sheet.cell(row=13 + row_offset, column=4 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(4, 7):
                    cell = sheet.cell(row=13 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)

                # Colocar "Lote:" en la celda G13
                merged_cell = sheet.cell(row=13 + row_offset, column=7 + col_offset)
                merged_cell.value = 'Lote:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño

                # Unir celdas H13:J13 y Colocar la raya
                sheet.merge_cells(start_row=13 + row_offset, start_column=8 + col_offset, end_row=13 + row_offset, end_column=10 + col_offset)
                merged_cell = sheet.cell(row=13 + row_offset, column=8 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(8, 11):
                    cell = sheet.cell(row=13 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)


                # Unir celdas L13:O13 y Colocar Fecha de vencimiento
                sheet.merge_cells(start_row=13 + row_offset, start_column=12 + col_offset, end_row=13 + row_offset, end_column=15 + col_offset)
                merged_cell = sheet.cell(row=13 + row_offset, column=12 + col_offset)
                merged_cell.value = 'Fecha de vencimiento:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño

                 # Unir celdas P13:Q13 y Colocar la raya
                sheet.merge_cells(start_row=13 + row_offset, start_column=16 + col_offset, end_row=13 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=13 + row_offset, column=16 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=11)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(16, 18):
                    cell = sheet.cell(row=13 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)


                # Unir celdas B14:C14 y Colocar Laboratorio
                sheet.merge_cells(start_row=14 + row_offset, start_column=2 + col_offset, end_row=14 + row_offset, end_column=3 + col_offset)
                merged_cell = sheet.cell(row=14 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Laboratorio:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño

               # Unir celdas D14:J14 y colocar "El laboratorio"
                sheet.merge_cells(start_row=14 + row_offset, start_column=4 + col_offset, end_row=14 + row_offset, end_column=10 + col_offset)
                merged_cell = sheet.cell(row=14 + row_offset, column=4 + col_offset)
                if lab:
                    merged_cell.value = lab.name
                else:
                    merged_cell.value = ''
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(4, 11):
                    cell = sheet.cell(row=14 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)

                # Unir celdas L14:O14 y Colocar ID Hermes
                sheet.merge_cells(start_row=14 + row_offset, start_column=12 + col_offset, end_row=14 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=14 + row_offset, column=12 + col_offset)
                merged_cell.value = 'ID Hermes:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño

                 # Unir celdas N14:Q14 y Colocar la raya y el id de hermes
                sheet.merge_cells(start_row=14 + row_offset, start_column=14 + col_offset, end_row=14 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=14 + row_offset, column=14 + col_offset)
                merged_cell.value = id_hermes
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=11)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(14, 18):
                    cell = sheet.cell(row=14 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)
                
                # Unir celdas B15:F15 y Colocar Teléfono de Emergencia UNAL
                sheet.merge_cells(start_row=15 + row_offset, start_column=2 + col_offset, end_row=15 + row_offset, end_column=6 + col_offset)
                merged_cell = sheet.cell(row=15 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Teléfono de Emergencia UNAL:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño

                # Unir celdas G15:J15 y colocar la raya y el teléfono de emergencia
                sheet.merge_cells(start_row=15 + row_offset, start_column=7 + col_offset, end_row=15 + row_offset, end_column=10 + col_offset)
                merged_cell = sheet.cell(row=15 + row_offset, column=7 + col_offset)
                if numero_emergencia:
                    merged_cell.value = numero_emergencia
                else:
                    merged_cell.value = ''
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(7, 11):
                    cell = sheet.cell(row=15 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)
                
                # Unir celdas L15:O15 y Sede
                # sheet.merge_cells(start_row=15 + row_offset, start_column=12 + col_offset, end_row=15 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=15 + row_offset, column=12 + col_offset)
                merged_cell.value = 'Sede:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño

                # Unir celdas N15:Q15 y Colocar la raya 
                sheet.merge_cells(start_row=15 + row_offset, start_column=13 + col_offset, end_row=15 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=15 + row_offset, column=13 + col_offset)
                merged_cell.value='Manizales'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=11)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(13, 18):
                    cell = sheet.cell(row=15 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)
                
                # Unir celdas B16:C16 y Colocar Proveedor
                sheet.merge_cells(start_row=16 + row_offset, start_column=2 + col_offset, end_row=16 + row_offset, end_column=3 + col_offset)
                merged_cell = sheet.cell(row=16 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Proveedor:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño
                
                # Unir celdas D16:Q16 y Colocar la raya 
                sheet.merge_cells(start_row=16 + row_offset, start_column=4 + col_offset, end_row=16 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=16 + row_offset, column=4 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=11)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(4, 18):
                    cell = sheet.cell(row=16 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)

                # Unir celdas B17:C17 y Colocar Dirección
                sheet.merge_cells(start_row=17 + row_offset, start_column=2 + col_offset, end_row=17 + row_offset, end_column=3 + col_offset)
                merged_cell = sheet.cell(row=17 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Dirección:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12)  # Fuente y tamaño
            
                # Unir celdas D17:J17 y Colocar la raya 
                sheet.merge_cells(start_row=17 + row_offset, start_column=4 + col_offset, end_row=17 + row_offset, end_column=10 + col_offset)
                merged_cell = sheet.cell(row=17 + row_offset, column=4 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=11)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(4, 11):
                    cell = sheet.cell(row=17 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)
                
                # Unir celdas L17:O17 y Telefóno
                # sheet.merge_cells(start_row=17 + row_offset, start_column=12 + col_offset, end_row=17 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=17 + row_offset, column=12 + col_offset)
                merged_cell.value = 'Teléfono:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=11)  # Fuente y tamaño

                # Unir celdas N15:Q15 y Colocar la raya 
                sheet.merge_cells(start_row=17 + row_offset, start_column=13 + col_offset, end_row=17 + row_offset, end_column=17 + col_offset)
                merged_cell = sheet.cell(row=17 + row_offset, column=13 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=11)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(13, 18):
                    cell = sheet.cell(row=17 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)
         # Configurar márgenes estrechos sin encabezado ni pie de página

        sheet.page_margins = PageMargins(left=1/2.54, right=1/2.54, top=1/2.54, bottom=1/2.54, header=0, footer=0)

        # Configurar la orientación de la página
        # sheet.page_setup = PrintPageSetup(orientation='landscape')

        # Guardar el archivo en la memoria y preparar la respuesta HTTP
        file_name = get_file_name()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        workbook.save(response)
        return response
    

# ----------------------- #
# Generar Etiqueta Mediana #
class GenerateMediumLabel(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
       
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def get(self, request, *args, **kwargs):
        def get_file_name():
            nombre_sustancia = request.GET.get('substance_name')
            now = datetime.now()
            return f'Etiqueta_mediana_{nombre_sustancia}_{now.strftime("%d%m%Y%H%M%S")}.xlsx'
        
        # Extraer parámetros de la URL
        nombre_sustancia = request.GET.get('substance_name')
        id_sustancia = request.GET.get('substance_id')
        
        numero_etiquetas = int(request.GET.get('label_number'))  # Convertir a entero
        id_hermes = request.GET.get('hermes_id')
        lab = request.GET.get('lab')
        if lab =='':
            lab=''
        else:
            lab=get_object_or_404(Laboratorios, id=lab)
        numero_emergencia = request.GET.get('e_number')        

        # Verificar si el número de etiquetas es superior a 56
        if numero_etiquetas > 8:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Mensaje de error
            error_message = "¡Upps algo ocurrió mal!\nEl número máximo de etiquetas debe ser menor o igual a 8, revisa y genera nuevamente tus etiquetas."

            # Ajustar la celda para el mensaje
            cell = sheet.cell(row=1, column=1)
            cell.value = error_message
            cell.font = Font(name='Arial', size=10, bold=True, color='FF0000')  # Letras llamativas, tamaño más pequeño
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Ajuste de texto y nueva línea
            sheet.merge_cells('A1:E10')  # Unir celdas para el mensaje, aumentando la altura
            sheet.row_dimensions[1].height = 40  # Ajustar la altura de la fila
            sheet.column_dimensions[get_column_letter(1)].width = 50  # Ajustar el ancho de la columna
            
            # Guardar el archivo en la memoria y preparar la respuesta HTTP
            file_name = get_file_name()
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={file_name}'
            workbook.save(response)
            return response

        sustancia = get_object_or_404(Sustancias, id=id_sustancia)
        cas = sustancia.cas
        if sustancia.warning:
            advertencia = sustancia.warning.name
        else:
            advertencia=''
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Configurar la orientación de la página
        sheet.page_setup = PrintPageSetup(orientation='landscape')

        # Definir un relleno con fondo blanco
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # Crear un borde de puntos
        thin_border = Side(style='thin')

        # Calcular número de líneas necesarias
        num_lines = math.ceil(numero_etiquetas / 2)
        

        for line in range(num_lines):
            # Calcular el offset de fila para cada línea
            row_offset = line * 18  # 18 filas por etiqueta (17 etiqueta + 1 separación)

            for i in range(2):
                index = line * 2 + i
                if index >= numero_etiquetas:
                    break

                # Calcular el offset de columna para cada etiqueta
                col_offset = i * 15
                
                # Aplicar el fondo blanco a las celdas de la etiqueta actual
                for row in range(1, 18):  # Filas 1 a 17
                    for col in range(1, 15):  # Columnas A a N
                        cell = sheet.cell(row=row + row_offset, column=col + col_offset)
                        cell.fill = white_fill

                # Anchos de columna para cada etiqueta
                sheet.column_dimensions[get_column_letter(1 + col_offset)].width = 0.60     # Ancho A = 0.39
                sheet.column_dimensions[get_column_letter(2 + col_offset)].width = 6.90     # Ancho B= 6.27
                sheet.column_dimensions[get_column_letter(3 + col_offset)].width = 9.10     # Ancho C = 8.45 
                sheet.column_dimensions[get_column_letter(4 + col_offset)].width = 0.50     # Ancho D = 0.28
                sheet.column_dimensions[get_column_letter(5 + col_offset)].width = 9.10     # Ancho E = 8.45 
                sheet.column_dimensions[get_column_letter(6 + col_offset)].width = 0.50     # Ancho F = 0.28
                sheet.column_dimensions[get_column_letter(7 + col_offset)].width = 4.55     # Ancho G = 3.91
                sheet.column_dimensions[get_column_letter(8 + col_offset)].width = 4.55     # Ancho H = 3.91
                sheet.column_dimensions[get_column_letter(9 + col_offset)].width = 0.50     # Ancho I = 0.28
                sheet.column_dimensions[get_column_letter(10 + col_offset)].width = 9.10    # Ancho J = 8.45
                sheet.column_dimensions[get_column_letter(11 + col_offset)].width = 0.50    # Ancho K = 0.28
                sheet.column_dimensions[get_column_letter(12 + col_offset)].width = 9.10    # Ancho L = 8.45
                sheet.column_dimensions[get_column_letter(13 + col_offset)].width = 6.90    # Ancho M = 6.27
                sheet.column_dimensions[get_column_letter(14 + col_offset)].width = 0.60    # Ancho N = 0.39

                # Separador para la siguiente etiqueta a la derecha
                
                
                sheet.column_dimensions[get_column_letter(15 + col_offset)].width = 0.50    # Ancho O = 0.28
                
                # Ajustar las alturas de las filas
                sheet.row_dimensions[1 + row_offset].height = 5.8       # Alto 1 = 5.8
                sheet.row_dimensions[2 + row_offset].height = 22.8      # Alto 2 = 22.8
                sheet.row_dimensions[3 + row_offset].height = 14.5      # Alto 3 = 14.5   
                sheet.row_dimensions[4 + row_offset].height = 12.5      # Alto 4 = 12.4 
                sheet.row_dimensions[5 + row_offset].height = 2.8       # Alto 5 = 2.8
                sheet.row_dimensions[6 + row_offset].height = 51        # Alto 6 = 51  
                sheet.row_dimensions[7 + row_offset].height = 2.8       # Alto 7 = 2.8  
                sheet.row_dimensions[8 + row_offset].height = 11.3      # Alto 8 = 11.3 
                sheet.row_dimensions[9 + row_offset].height = 21.5      # Alto 9 = 21.5
                sheet.row_dimensions[10 + row_offset].height = 14.5     # Alto 10 = 14.5
                # sheet.row_dimensions[11 + row_offset].height = 119      # Alto 11 = 119
                sheet.row_dimensions[12+ row_offset].height = 14.3      # Alto 12 = 14.3
                sheet.row_dimensions[13+ row_offset].height = 14.3      # Alto 13 = 14.3
                sheet.row_dimensions[14+ row_offset].height = 14.3      # Alto 14 = 14.3
                sheet.row_dimensions[15+ row_offset].height = 14.3      # Alto 15 = 14.3
                sheet.row_dimensions[16+ row_offset].height = 14.3      # Alto 16 = 14.3
                sheet.row_dimensions[17+ row_offset].height = 5.8       # Alto 17 = 5.8
                

                # Separador para la siguiente etiqueta abajo
                sheet.row_dimensions[18+ row_offset].height = 5.8       # Alto 18 = 5.8           
                
                # Aplicar bordes manualmente para formar el rectángulo
                # A1: Borde izquierdo y borde arriba
                sheet.cell(row=1 + row_offset, column=1 + col_offset).border = Border(left=thin_border, top=thin_border)
               # A2:A16 Borde izquierdo
                for row in range(2, 17):
                    sheet.cell(row=row + row_offset, column=1 + col_offset).border = Border(left=thin_border)

                # A17 Borde izquierdo y borde abajo
                sheet.cell(row=17 + row_offset, column=1 + col_offset).border = Border(left=thin_border, bottom=thin_border)

                # B1:M1 Borde arriba
                for col in range(2, 14):
                    sheet.cell(row=1 + row_offset, column=col + col_offset).border = Border(top=thin_border)

                # N1: Borde arriba, borde derecho
                sheet.cell(row=1 + row_offset, column=14 + col_offset).border = Border(top=thin_border, right=thin_border)

                # B17:M17 Borde abajo
                for col in range(2, 14):
                    sheet.cell(row=17 + row_offset, column=col + col_offset).border = Border(bottom=thin_border)

                # Nl7 Borde abajo, borde derecho
                sheet.cell(row=17 + row_offset, column=14 + col_offset).border = Border(bottom=thin_border, right=thin_border)

                # N2:N16 Borde derecho
                for row in range(2, 17):
                    sheet.cell(row=row + row_offset, column=14 + col_offset).border = Border(right=thin_border)

                #### Unir celdas B2:M2 y colocar "Nombre de la sustancia" ####
                sheet.merge_cells(start_row=2 + row_offset, start_column=2 + col_offset, end_row=2 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=2 + row_offset, column=2 + col_offset)
                merged_cell.value = nombre_sustancia.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=18, bold=True)  # Fuente, tamaño y negrita

                
                # Unir celdas B3:G3 y colocar "CAS"
                sheet.merge_cells(start_row=3 + row_offset, start_column=2 + col_offset, end_row=3 + row_offset, end_column=7 + col_offset)
                merged_cell = sheet.cell(row=3 + row_offset, column=2 + col_offset)
                merged_cell.value = 'CAS'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=11, bold=True)  # Fuente y tamaño

                # Unir celdas H3:M3 y colocar CAS de la sustancia
                sheet.merge_cells(start_row=3 + row_offset, start_column=8 + col_offset, end_row=3 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=3 + row_offset, column=8 + col_offset)
                merged_cell.value = cas.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=11, bold=True)  # Fuente y tamaño

                # Unir celdas B4:M4 y colocar "Pictogramas"
                sheet.merge_cells(start_row=4 + row_offset, start_column=2 + col_offset, end_row=4 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=4 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Pictogramas:'  
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=10, underline='single')  # Fuente, tamaño y negrita

                # C6: Agregar el pictograma1 si existe
                if sustancia and sustancia.pictogram_clp1:
                    fila_picto=6+row_offset
                    pictogram_path = sustancia.pictogram_clp1.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 68     # Ajuste Altura 1.8
                    img.width = 68      # Ajuste Ancho 1.8
                    img.anchor = f'{get_column_letter(3 + col_offset)}{fila_picto}'
                    sheet.add_image(img)

                # E6: Agregar el pictograma2 si existe
                if sustancia and sustancia.pictogram_clp2:
                    fila_picto=6+row_offset
                    pictogram_path = sustancia.pictogram_clp2.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 68     # Ajuste Altura 1.8
                    img.width = 68      # Ajuste Ancho 1.8
                    img.anchor = f'{get_column_letter(5 + col_offset)}{fila_picto}'
                    sheet.add_image(img)
                
                # G6: Agregar el pictograma3 si existe
                if sustancia and sustancia.pictogram_clp3:
                    fila_picto=6+row_offset
                    pictogram_path = sustancia.pictogram_clp3.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 68     # Ajuste Altura 1.8
                    img.width = 68      # Ajuste Ancho 1.8
                    img.anchor = f'{get_column_letter(7 + col_offset)}{fila_picto}'
                    sheet.add_image(img)
                
                # J6: Agregar el pictograma4 si existe
                if sustancia and sustancia.pictogram_clp4:
                    fila_picto=6+row_offset
                    pictogram_path = sustancia.pictogram_clp4.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 68     # Ajuste Altura 1.8
                    img.width = 68      # Ajuste Ancho 1.8
                    img.anchor = f'{get_column_letter(10 + col_offset)}{fila_picto}'
                    sheet.add_image(img)
                
                # L6: Agregar el pictograma5 si existe
                if sustancia and sustancia.pictogram_clp5:
                    fila_picto=6+row_offset
                    pictogram_path = sustancia.pictogram_clp5.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 68     # Ajuste Altura 1.8
                    img.width = 68      # Ajuste Ancho 1.8
                    img.anchor = f'{get_column_letter(12 + col_offset)}{fila_picto}'
                    sheet.add_image(img)

                # Unir celdas B8:M8 y colocar "Palabra de Advertencia"
                sheet.merge_cells(start_row=8 + row_offset, start_column=2 + col_offset, end_row=8 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=8 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Palabra de Advertencia:'  
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=9,)  # Fuente, tamaño y negrita

                # Unir celdas B9:M9 y colocar "La Palabra de Advertencia"
                sheet.merge_cells(start_row=9 + row_offset, start_column=2 + col_offset, end_row=9 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=9 + row_offset, column=2 + col_offset)
                merged_cell.value = advertencia.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=16, bold=True, color="FF0000")  # Fuente, tamaño y negrita
                
                
                # Unir celdas B10:G10 y Colocar Frases H
                sheet.merge_cells(start_row=10 + row_offset, start_column=2 + col_offset, end_row=10 + row_offset, end_column=7 + col_offset)
                merged_cell = sheet.cell(row=10 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Frases H'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12, bold=True)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(2, 8):
                    cell = sheet.cell(row=10 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)
                
                # Unir celdas H10:M10 y Colocar Frases P
                sheet.merge_cells(start_row=10 + row_offset, start_column=8 + col_offset, end_row=10 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=10 + row_offset, column=8 + col_offset)
                merged_cell.value = 'Frases P'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12, bold=True)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(8, 14):
                    cell = sheet.cell(row=10 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)

                # Organizar Frases H
                frases_h = '\n'.join(sustancia.phrase_h.values_list('name', flat=True)) if sustancia.phrase_h.exists() else '-'

                # Unir celdas B11:G11 y Colocar las Frases H
                sheet.merge_cells(start_row=11 + row_offset, start_column=2 + col_offset, end_row=11 + row_offset, end_column=7 + col_offset)
                merged_cell = sheet.cell(row=11 + row_offset, column=2 + col_offset)
                merged_cell.value = frases_h
                merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                merged_cell.font = Font(name='Ancizar Sans', size=7)  # Fuente y tamaño

                # Configurar bordes superior e inferior
                for col in range(2, 8):
                    cell = sheet.cell(row=11 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)
                    

                # Organizar Frases P
                frases_p = '\n'.join(sustancia.phrase_p.values_list('name', flat=True)) if sustancia.phrase_p.exists() else '-'

                # Unir celdas H11:M11 y Colocar las Frases P
                sheet.merge_cells(start_row=11 + row_offset, start_column=8 + col_offset, end_row=11 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=11 + row_offset, column=8 + col_offset)
                merged_cell.value = frases_p
                merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)  # Ajuste de texto
                merged_cell.font = Font(name='Ancizar Sans', size=7)  # Fuente y tamaño

                # Configurar bordes superior e inferior
                for col in range(8, 14):
                    cell = sheet.cell(row=11 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)
                
                # Ajustar altura de las filas según el contenido
                if not frases_h=='-' or not frases_p=='-':
                    max_lines_h = max(len(frases_h.split('\n')), 1)  # Calcular el número de líneas para frases H
                    max_lines_p = max(len(frases_p.split('\n')), 1)  # Calcular el número de líneas para frases P
                    max_lines = max(max_lines_h, max_lines_p)  # Obtener el mayor número de líneas entre H y P
                    line_height = 9  # Altura estimada por línea (puedes ajustar este valor según la fuente y el tamaño)
                    total_height = max_lines * line_height * 2

                    # Ajustar las alturas de la fila 11
                    sheet.row_dimensions[11 + row_offset].height = total_height

                # Colocar "Cantidad" en la celda B12
                merged_cell = sheet.cell(row=12 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Cantidad:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Colocar "" en la celda C12
                merged_cell = sheet.cell(row=12 + row_offset, column=3 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                merged_cell.border = Border(bottom=thin_border)# Configurar borde inferior

                # Colocar "Lote:" en la celda E12
                merged_cell = sheet.cell(row=12 + row_offset, column=5 + col_offset)
                merged_cell.value = 'Lote:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas F12:H12 y colocar ""
                sheet.merge_cells(start_row=12 + row_offset, start_column=6 + col_offset, end_row=12 + row_offset, end_column=8 + col_offset)
                merged_cell = sheet.cell(row=12 + row_offset, column=6 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(6, 9):
                    cell = sheet.cell(row=12 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)

                # Colocar "Vencimiento:" en la celda J12
                merged_cell = sheet.cell(row=12 + row_offset, column=10 + col_offset)
                merged_cell.value = 'Vencimiento:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas I12:M12 y colocar ""
                sheet.merge_cells(start_row=12 + row_offset, start_column=11 + col_offset, end_row=12 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=12 + row_offset, column=11 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(11, 14):
                    cell = sheet.cell(row=12 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)


                # Colocar "Laboratorio" en la celda B13
                merged_cell = sheet.cell(row=13 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Laboratorio:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

               # Unir celdas C13:H13 y colocar "El laboratorio"
                sheet.merge_cells(start_row=13 + row_offset, start_column=3 + col_offset, end_row=13 + row_offset, end_column=8 + col_offset)
                merged_cell = sheet.cell(row=13 + row_offset, column=3 + col_offset)
                if lab:
                    merged_cell.value = lab.name
                else:
                    merged_cell.value = ''
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(3, 9):
                    cell = sheet.cell(row=13 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)

                # Colocar "ID Hermes:" en la celda J13
                merged_cell = sheet.cell(row=13 + row_offset, column=10 + col_offset)
                merged_cell.value = 'ID Hermes:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas I13:M13 y colocar "El ID de Hermes"
                sheet.merge_cells(start_row=13 + row_offset, start_column=11 + col_offset, end_row=13 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=13 + row_offset, column=11 + col_offset)
                merged_cell.value = id_hermes
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(11, 14):
                    cell = sheet.cell(row=13 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)
                
                # Unir celdas B14:C14 y colocar "Teléfono Emergencia UNAL"
                sheet.merge_cells(start_row=14 + row_offset, start_column=2 + col_offset, end_row=14 + row_offset, end_column=3 + col_offset)
                merged_cell = sheet.cell(row=14 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Teléfono Emergencia UNAL:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaños

                # Unir celdas D13:M13 y colocar ""
                sheet.merge_cells(start_row=14 + row_offset, start_column=4 + col_offset, end_row=14 + row_offset, end_column=8 + col_offset)
                merged_cell = sheet.cell(row=14 + row_offset, column=4 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(4, 9):
                    cell = sheet.cell(row=14 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)
                
                # Colocar "Sede:" en la celda J14
                merged_cell = sheet.cell(row=14 + row_offset, column=10 + col_offset)
                merged_cell.value = 'Sede:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas I14:M14 y colocar por defecto "Manizales"
                sheet.merge_cells(start_row=14 + row_offset, start_column=11 + col_offset, end_row=14 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=14 + row_offset, column=11 + col_offset)
                merged_cell.value = 'Manizales'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(11, 14):
                    cell = sheet.cell(row=14 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)
                
                # Colocar "Proveedor" en la celda B15
                merged_cell = sheet.cell(row=15 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Proveedor:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas C15:M15 y colocar por defecto ""
                sheet.merge_cells(start_row=15 + row_offset, start_column=3 + col_offset, end_row=15 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=15 + row_offset, column=3   + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(3, 14):
                    cell = sheet.cell(row=15 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)

                # Colocar "Dirección" en la celda B16
                merged_cell = sheet.cell(row=16 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Dirección:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas C16:H16 y colocar por defecto ""
                sheet.merge_cells(start_row=16 + row_offset, start_column=3 + col_offset, end_row=16 + row_offset, end_column=8 + col_offset)
                merged_cell = sheet.cell(row=16 + row_offset, column=3   + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(3, 9):
                    cell = sheet.cell(row=16 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)
                
                # Colocar "Teléfono:" en la celda J16
                merged_cell = sheet.cell(row=16 + row_offset, column=10 + col_offset)
                merged_cell.value = 'Teléfono:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas I16:M16 y colocar por defecto ""
                sheet.merge_cells(start_row=16 + row_offset, start_column=11 + col_offset, end_row=16 + row_offset, end_column=13 + col_offset)
                merged_cell = sheet.cell(row=16 + row_offset, column=11 + col_offset)
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Configurar borde inferior
                for col in range(11, 14):
                    cell = sheet.cell(row=16 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)
         # Configurar márgenes estrechos sin encabezado ni pie de página

        sheet.page_margins = PageMargins(left=0.64/2.54, right=0.64/2.54, top=0.64/2.54, bottom=0.64/2.54, header=0, footer=0)

        # Configurar la orientación de la página
        sheet.page_setup = PrintPageSetup(orientation='landscape')

        # Guardar el archivo en la memoria y preparar la respuesta HTTP
        file_name = get_file_name()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        workbook.save(response)
        return response




# ----------------------- #
# Generar Etiqueta Pequeña #
class GenerateSmallLabel(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
       
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def get(self, request, *args, **kwargs):
        def get_file_name():
            nombre_sustancia = request.GET.get('substance_name')
            now = datetime.now()
            return f'Etiqueta_pequeña_{nombre_sustancia}_{now.strftime("%d%m%Y%H%M%S")}.xlsx'
        
        # Extraer parámetros de la URL
        nombre_sustancia = request.GET.get('substance_name')
        id_sustancia = request.GET.get('substance_id')
        
        numero_etiquetas = int(request.GET.get('label_number'))  # Convertir a entero
        id_hermes = request.GET.get('hermes_id')
        lab = request.GET.get('lab')
        if lab =='':
            lab=''
        else:
            lab=get_object_or_404(Laboratorios, id=lab)
        numero_emergencia = request.GET.get('e_number')        

        # Verificar si el número de etiquetas es superior a 56
        if numero_etiquetas > 24:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Mensaje de error
            error_message = "¡Upps algo ocurrió mal!\nEl número máximo de etiquetas debe ser menor o igual a 24, revisa y genera nuevamente tus etiquetas."

            # Ajustar la celda para el mensaje
            cell = sheet.cell(row=1, column=1)
            cell.value = error_message
            cell.font = Font(name='Arial', size=10, bold=True, color='FF0000')  # Letras llamativas, tamaño más pequeño
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Ajuste de texto y nueva línea
            sheet.merge_cells('A1:E10')  # Unir celdas para el mensaje, aumentando la altura
            sheet.row_dimensions[1].height = 40  # Ajustar la altura de la fila
            sheet.column_dimensions[get_column_letter(1)].width = 50  # Ajustar el ancho de la columna
            
            # Guardar el archivo en la memoria y preparar la respuesta HTTP
            file_name = get_file_name()
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={file_name}'
            workbook.save(response)
            return response

        sustancia = get_object_or_404(Sustancias, id=id_sustancia)
        cas = sustancia.cas
        if sustancia.warning:
            advertencia = sustancia.warning.name
        else:
            advertencia=''
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Configurar la orientación de la página
        sheet.page_setup = PrintPageSetup(orientation='landscape')

        # Definir un relleno con fondo blanco
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # Crear un borde de puntos
        thin_border = Side(style='thin')

        # Calcular número de líneas necesarias
        num_lines = math.ceil(numero_etiquetas / 2)
        

        for line in range(num_lines):
            # Calcular el offset de fila para cada línea
            row_offset = line * 12  # 12 filas por etiqueta (11 etiquetas + 1 separación)

            for i in range(2):
                index = line * 2 + i
                if index >= numero_etiquetas:
                    break

                # Calcular el offset de columna para cada etiqueta
                col_offset = i * 14
                
                # Aplicar el fondo blanco a las celdas de la etiqueta actual
                for row in range(1, 12):  # Filas 1 a 11
                    for col in range(1, 14):  # Columnas A a M
                        cell = sheet.cell(row=row + row_offset, column=col + col_offset)
                        cell.fill = white_fill

                # Anchos de columna para cada etiqueta
                sheet.column_dimensions[get_column_letter(1 + col_offset)].width = 0.98     # Ancho = 0.61
                sheet.column_dimensions[get_column_letter(2 + col_offset)].width = 4.65     # Ancho = 4
                sheet.column_dimensions[get_column_letter(3 + col_offset)].width = 3        # Ancho = 2.36 
                sheet.column_dimensions[get_column_letter(4 + col_offset)].width = 1.2      # Ancho 0.72
                sheet.column_dimensions[get_column_letter(5 + col_offset)].width = 4.10      # Ancho 3.45 
                sheet.column_dimensions[get_column_letter(6 + col_offset)].width = 1.68     # Ancho = 1
                sheet.column_dimensions[get_column_letter(7 + col_offset)].width = 1.55     # Ancho = 0.94
                sheet.column_dimensions[get_column_letter(8 + col_offset)].width = 4.55      # Ancho = 3.91
                sheet.column_dimensions[get_column_letter(9 + col_offset)].width = 7.45      # Ancho = 6.82
                sheet.column_dimensions[get_column_letter(10 + col_offset)].width = 7.45     # Ancho = 6.82
                sheet.column_dimensions[get_column_letter(11 + col_offset)].width = 7.45    # Ancho = 6.82
                sheet.column_dimensions[get_column_letter(12 + col_offset)].width = 7.45    # Ancho = 6.82
                sheet.column_dimensions[get_column_letter(13 + col_offset)].width = 0.98    # Ancho = 0.61

                # Separador para la siguiente etiqueta a la derecha
                
                
                sheet.column_dimensions[get_column_letter(14 + col_offset)].width = 1.30    # Ancho = 0.78
                

                # # Ajustar las alturas de las filas
                

                
                
                # Ajustar las alturas de las filas
                sheet.row_dimensions[1 + row_offset].height = 5.8   # Alto 5.8
                sheet.row_dimensions[2 + row_offset].height = 13  # Alto 13
                sheet.row_dimensions[3 + row_offset].height = 9     # Alto 9   
                sheet.row_dimensions[4 + row_offset].height = 4.5   # Alto 4.5 
                sheet.row_dimensions[5 + row_offset].height = 16.5  # Alto 16.5
                sheet.row_dimensions[6 + row_offset].height = 20    # Alto 20  
                sheet.row_dimensions[7 + row_offset].height = 40    # Alto 40  
                sheet.row_dimensions[8 + row_offset].height = 5.3   # Alto 5.3 
                sheet.row_dimensions[9 + row_offset].height = 14.5  # Alto 14.5
                sheet.row_dimensions[10 + row_offset].height = 11 # Alto 11
                sheet.row_dimensions[11 + row_offset].height = 5.8  # Alto 5.8
                # Separador para la siguiente etiqueta abajo
                sheet.row_dimensions[12+ row_offset].height = 15   # Alto 15             
                
                # Aplicar bordes manualmente para formar el rectángulo
                # A1: Borde izquierdo y borde arriba
                sheet.cell(row=1 + row_offset, column=1 + col_offset).border = Border(left=thin_border, top=thin_border)

                # A2:A10 Borde izquierdo
                for row in range(2, 11):
                    sheet.cell(row=row + row_offset, column=1 + col_offset).border = Border(left=thin_border)

                # A11 Borde izquierdo y borde abajo
                sheet.cell(row=11 + row_offset, column=1 + col_offset).border = Border(left=thin_border, bottom=thin_border)

                # B1:L1 Borde arriba
                for col in range(2, 13):
                    sheet.cell(row=1 + row_offset, column=col + col_offset).border = Border(top=thin_border)

                # M1: Borde arriba, borde derecho
                sheet.cell(row=1 + row_offset, column=13 + col_offset).border = Border(top=thin_border, right=thin_border)

                # B11:L11 Borde abajo
                for col in range(2, 13):
                    sheet.cell(row=11 + row_offset, column=col + col_offset).border = Border(bottom=thin_border)

                # Mll Borde abajo, borde derecho
                sheet.cell(row=11 + row_offset, column=13 + col_offset).border = Border(bottom=thin_border, right=thin_border)

                # M2:M11 Borde derecho
                for row in range(2, 11):
                    sheet.cell(row=row + row_offset, column=13 + col_offset).border = Border(right=thin_border)

                #### Unir celdas B2:L2 y colocar "Nombre de la sustancia" ####
                sheet.merge_cells(start_row=2 + row_offset, start_column=2 + col_offset, end_row=2 + row_offset, end_column=12 + col_offset)
                merged_cell = sheet.cell(row=2 + row_offset, column=2 + col_offset)
                merged_cell.value = nombre_sustancia.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12, bold=True)  # Fuente, tamaño y negrita

                # Unir celdas K3:L3 y  colocar teléfono"
                sheet.merge_cells(start_row=3 + row_offset, start_column=11 + col_offset, end_row=3 + row_offset, end_column=12 + col_offset)
                merged_cell = sheet.cell(row=3 + row_offset, column=11 + col_offset)
                merged_cell.value =numero_emergencia
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Aplicar borde inferior a las celdas combinadas K3:L3
                for col in range(11, 13):
                    cell = sheet.cell(row=3 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)

                # Unir celdas I3:J3 y COLOCAR teléfono emergencia
                sheet.merge_cells(start_row=3 + row_offset, start_column=9 + col_offset, end_row=3 + row_offset, end_column=10 + col_offset)
                merged_cell = sheet.cell(row=3 + row_offset, column=9 + col_offset)
                merged_cell.value = 'Teléfono Emergencia.'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8, bold=True)  # Fuente y tamaño

                # Unir celdas B3:C3 y colocar "CAS"
                sheet.merge_cells(start_row=3 + row_offset, start_column=2 + col_offset, end_row=3 + row_offset, end_column=3 + col_offset)
                merged_cell = sheet.cell(row=3 + row_offset, column=2 + col_offset)
                merged_cell.value = 'CAS'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas D3:H3 y colocar CAS de la sustancia
                sheet.merge_cells(start_row=3 + row_offset, start_column=4 + col_offset, end_row=3 + row_offset, end_column=8 + col_offset)
                merged_cell = sheet.cell(row=3 + row_offset, column=4 + col_offset)
                merged_cell.value = cas.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas B4:B7, girar 90° a la izquierda y colocar palabra de advertencia en mayúscula y negrita
                sheet.merge_cells(start_row=4 + row_offset, start_column=2 + col_offset, end_row=7 + row_offset, end_column=2 + col_offset)
                merged_cell = sheet.cell(row=4 + row_offset, column=2 + col_offset)
                merged_cell.value = advertencia.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True, text_rotation=90)  # Girar texto 90°
                merged_cell.font = Font(name='Ancizar Sans', size=12, bold=True, color="FF0000")  # Fuente, tamaño, negrita, color

                # C4: Agregar el pictograma1 si existe
                if sustancia and sustancia.pictogram_clp1:
                    fila_picto=4+row_offset
                    pictogram_path = sustancia.pictogram_clp1.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 53     # Ajuste Altura
                    img.width = 53      # Ajuste Ancho
                    img.anchor = f'{get_column_letter(3 + col_offset)}{fila_picto}'
                    sheet.add_image(img)
                
                # C7: Agregar el pictograma2 si existe
                if sustancia and sustancia.pictogram_clp2:
                    fila_picto=7+row_offset
                    pictogram_path = sustancia.pictogram_clp2.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 53     # Ajuste Altura
                    img.width = 53      # Ajuste Ancho
                    img.anchor = f'{get_column_letter(3 + col_offset)}{fila_picto}'
                    sheet.add_image(img)

                # E6: Agregar el pictograma3 si existe
                if sustancia and sustancia.pictogram_clp3:
                    fila_picto=6+row_offset
                    pictogram_path = sustancia.pictogram_clp3.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 53     # Ajuste Altura
                    img.width = 53      # Ajuste Ancho
                    img.anchor = f'{get_column_letter(5 + col_offset)}{fila_picto}'
                    sheet.add_image(img)

                # F4: Agregar el pictograma4 si existe
                if sustancia and sustancia.pictogram_clp4:
                    fila_picto=4+row_offset
                    pictogram_path = sustancia.pictogram_clp4.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 53     # Ajuste Altura
                    img.width = 53      # Ajuste Ancho
                    img.anchor = f'{get_column_letter(6 + col_offset)}{fila_picto}'
                    sheet.add_image(img)

                # F7: Agregar el pictograma5 si existe
                if sustancia and sustancia.pictogram_clp5:
                    fila_picto=7+row_offset
                    pictogram_path = sustancia.pictogram_clp5.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 53     # Ajuste Altura
                    img.width = 53      # Ajuste Ancho
                    img.anchor = f'{get_column_letter(6 + col_offset)}{fila_picto}'
                    sheet.add_image(img)
                
                # Unir celdas I5:J5 y Colocar Frases H
                sheet.merge_cells(start_row=5 + row_offset, start_column=9 + col_offset, end_row=5 + row_offset, end_column=10 + col_offset)
                merged_cell = sheet.cell(row=5 + row_offset, column=9 + col_offset)
                merged_cell.value = 'Frases H'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12, bold=True)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(9, 11):
                    cell = sheet.cell(row=5 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)
                
                # Unir celdas K5:L5 y Colocar Frases P
                sheet.merge_cells(start_row=5 + row_offset, start_column=11 + col_offset, end_row=5 + row_offset, end_column=12 + col_offset)
                merged_cell = sheet.cell(row=5 + row_offset, column=11 + col_offset)
                merged_cell.value = 'Frases P'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=12, bold=True)  # Fuente y tamaño
                # Configurar bordes superior e inferior                
                for col in range(11, 13):
                    cell = sheet.cell(row=5 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)


                # Organizar Frases H
                frases_h = '\n'.join(sustancia.phrase_h.values_list('name', flat=True)) if sustancia.phrase_h.exists() else '-'

                # Unir celdas I6:J7 y Colocar las Frases H
                sheet.merge_cells(start_row=6 + row_offset, start_column=9 + col_offset, end_row=7 + row_offset, end_column=10 + col_offset)
                merged_cell = sheet.cell(row=6 + row_offset, column=9 + col_offset)
                merged_cell.value = frases_h
                merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                merged_cell.font = Font(name='Ancizar Sans', size=6)  # Fuente y tamaño

                # Configurar bordes superior e inferior
                for col in range(9, 11):
                    cell = sheet.cell(row=6 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)
                    cell = sheet.cell(row=7 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)

                # Organizar Frases P
                frases_p = '\n'.join(sustancia.phrase_p.values_list('name', flat=True)) if sustancia.phrase_p.exists() else '-'

                # Unir celdas K6:L7 y Colocar las Frases P
                sheet.merge_cells(start_row=6 + row_offset, start_column=11 + col_offset, end_row=7 + row_offset, end_column=12 + col_offset)
                merged_cell = sheet.cell(row=6 + row_offset, column=11 + col_offset)
                merged_cell.value = frases_p
                merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)  # Ajuste de texto
                merged_cell.font = Font(name='Ancizar Sans', size=6)  # Fuente y tamaño

                # Configurar bordes superior e inferior
                for col in range(11, 13):
                    cell = sheet.cell(row=6 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)
                    cell = sheet.cell(row=7 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, top=thin_border, left=thin_border, right=thin_border)

                # Ajustar altura de las filas según el contenido
                max_lines_h = max(len(frases_h.split('\n')), 1)  # Calcular el número de líneas para frases H
                max_lines_p = max(len(frases_p.split('\n')), 1)  # Calcular el número de líneas para frases P
                max_lines = max(max_lines_h, max_lines_p)  # Obtener el mayor número de líneas entre H y P
                line_height = 12  # Altura estimada por línea (puedes ajustar este valor según la fuente y el tamaño)
                total_height = max_lines * line_height

                # Ajustar las alturas de la fila 7 teniendo presente que el tamaño de la fila 6 debe ser inamovible (20)
                sheet.row_dimensions[7 + row_offset].height = (total_height * 2) - 20


                # Linea Que divide los datos del reactivo con los adicionales                
                for col in range(2, 13):
                    cell = sheet.cell(row=8 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)

                # Unir celdas B9:C9 y colocar "Laboratorio:"
                sheet.merge_cells(start_row=9 + row_offset, start_column=2 + col_offset, end_row=9 + row_offset, end_column=3 + col_offset)
                merged_cell = sheet.cell(row=9 + row_offset, column=2 + col_offset)
                merged_cell.value = 'Laboratorio:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas D9:H9 y colocar "El Laboratorio:"
                sheet.merge_cells(start_row=9 + row_offset, start_column=4 + col_offset, end_row=9 + row_offset, end_column=8 + col_offset)
                merged_cell = sheet.cell(row=9 + row_offset, column=4 + col_offset)
                if lab:
                    merged_cell.value = lab.name
                else:
                    merged_cell.value = ''
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=6)  # Fuente y tamaño
                # Configurar borde inferior                
                for col in range(4, 9):
                    cell = sheet.cell(row=9 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border, right=thin_border)

                # Unir celdas B10:C10 y colocar "ID Hermes:"
                sheet.merge_cells(start_row=10 + row_offset, start_column=2 + col_offset, end_row=10 + row_offset, end_column=3 + col_offset)
                merged_cell = sheet.cell(row=10 + row_offset, column=2 + col_offset)
                merged_cell.value = 'ID Hermes:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas D10:E10 y colocar "El ID de hermes"
                sheet.merge_cells(start_row=10 + row_offset, start_column=4 + col_offset, end_row=10 + row_offset, end_column=5 + col_offset)
                merged_cell = sheet.cell(row=10 + row_offset, column=4 + col_offset)
                merged_cell.value = id_hermes
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=6)  # Fuente y tamaño
                # Configurar borde inferior                
                for col in range(4, 6):
                    cell = sheet.cell(row=10 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border)

                # Unir celdas F10:G10 y colocar "Sede:"
                sheet.merge_cells(start_row=10 + row_offset, start_column=6 + col_offset, end_row=10 + row_offset, end_column=7 + col_offset)
                merged_cell = sheet.cell(row=10 + row_offset, column=6 + col_offset)
                merged_cell.value = 'Sede:'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Colocar "Manizales" en la celda H10
                merged_cell = sheet.cell(row=10 + row_offset, column=8 + col_offset)
                merged_cell.value = 'Mzls'
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=6)  # Fuente y tamaño
                merged_cell.border = Border(bottom=thin_border,right=thin_border)# Configurar borde inferior

                # Colocar "Proveedor" en la celda I9
                merged_cell = sheet.cell(row=9 + row_offset, column=9 + col_offset)
                merged_cell.value = 'Proveedor: '
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas J9:L9 y colocar "Espacio para proveedor"
                sheet.merge_cells(start_row=9 + row_offset, start_column=10 + col_offset, end_row=9 + row_offset, end_column=12 + col_offset)
                merged_cell = sheet.cell(row=9 + row_offset, column=10 + col_offset)
                merged_cell.value = ''
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                # Configurar borde inferior                
                for col in range(10, 13):
                    cell = sheet.cell(row=9 + row_offset, column=col + col_offset)
                    cell.border = Border(bottom=thin_border,)
                
                # Colocar "Dirección" en la celda I10
                merged_cell = sheet.cell(row=10 + row_offset, column=9 + col_offset)
                merged_cell.value = 'Dirección: '
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Colocar "" en la celda J10
                merged_cell = sheet.cell(row=10 + row_offset, column=10 + col_offset)
                merged_cell.value = ''
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                merged_cell.border = Border(bottom=thin_border) # Configurar borde inferior

                # Colocar "Teléfono" en la celda K10
                merged_cell = sheet.cell(row=10 + row_offset, column=11 + col_offset)
                merged_cell.value = 'Teléfono: '
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Colocar "" en la celda L10
                merged_cell = sheet.cell(row=10 + row_offset, column=12 + col_offset)
                merged_cell.value = ''
                merged_cell.alignment = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño
                merged_cell.border = Border(bottom=thin_border) # Configurar borde inferior

        # Configurar la orientación de la página
        sheet.page_setup = PrintPageSetup(orientation='landscape')

        # Guardar el archivo en la memoria y preparar la respuesta HTTP
        file_name = get_file_name()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        workbook.save(response)
        return response


# ----------------------- #
# Generar Etiqueta Básica #
class GenerateBasicLabel(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
       
    @check_group_permission(groups_required=['ADMINISTRADOR, COORDINADOR, TECNICO'])
    def get(self, request, *args, **kwargs):
        def get_file_name():
            nombre_sustancia = request.GET.get('substance_name')
            now = datetime.now()
            return f'Etiqueta_básica_{nombre_sustancia}_{now.strftime("%d%m%Y%H%M%S")}.xlsx'
        
        # Extraer parámetros de la URL
        nombre_sustancia = request.GET.get('substance_name')
        id_sustancia = request.GET.get('substance_id')
        
        numero_etiquetas = int(request.GET.get('label_number'))  # Convertir a entero

        # Verificar si el número de etiquetas es superior a 56
        if numero_etiquetas > 56:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Mensaje de error
            # Mensaje de error
            error_message = "¡Upps algo ocurrió mal!\nEl número máximo de etiquetas debe ser menor o igual a 56, revisa y genera nuevamente tus etiquetas."

            # Ajustar la celda para el mensaje
            cell = sheet.cell(row=1, column=1)
            cell.value = error_message
            cell.font = Font(name='Arial', size=10, bold=True, color='FF0000')  # Letras llamativas, tamaño más pequeño
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Ajuste de texto y nueva línea
            sheet.merge_cells('A1:E10')  # Unir celdas para el mensaje, aumentando la altura
            sheet.row_dimensions[1].height = 40  # Ajustar la altura de la fila
            sheet.column_dimensions[get_column_letter(1)].width = 50  # Ajustar el ancho de la columna
            
            # Guardar el archivo en la memoria y preparar la respuesta HTTP
            file_name = get_file_name()
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={file_name}'
            workbook.save(response)
            return response

        sustancia = get_object_or_404(Sustancias, id=id_sustancia)
        cas = sustancia.cas
        if sustancia.warning:
            advertencia = sustancia.warning.name
        else:
            advertencia=''
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Definir un relleno con fondo blanco
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        # Crear un borde de puntos
        thin_border = Side(style='thin')

        # Calcular número de líneas necesarias
        num_lines = math.ceil(numero_etiquetas / 4)
        

        for line in range(num_lines):
            # Calcular el offset de fila para cada línea
            row_offset = line * 9  # 9 filas por etiqueta (8 etiquetas + 1 separación)

            for i in range(4):
                index = line * 4 + i
                if index >= numero_etiquetas:
                    break

                # Calcular el offset de columna para cada etiqueta
                col_offset = i * 7
                
                # Aplicar el fondo blanco a las celdas de la etiqueta actual
                for row in range(1, 9):  # Filas 1 a 8
                    for col in range(1, 7):  # Columnas A a F
                        cell = sheet.cell(row=row + row_offset, column=col + col_offset)
                        cell.fill = white_fill

                # Anchos de columna para cada etiqueta
                sheet.column_dimensions[get_column_letter(1 + col_offset)].width = 1.68  # Ancho = 1
                sheet.column_dimensions[get_column_letter(2 + col_offset)].width = 4.9   # Ancho = 4.27
                sheet.column_dimensions[get_column_letter(3 + col_offset)].width = 3     # Ancho = 2.36 
                sheet.column_dimensions[get_column_letter(4 + col_offset)].width = 2.90  # Ancho 2.27
                sheet.column_dimensions[get_column_letter(5 + col_offset)].width = 5.3   # Ancho 4.64 
                sheet.column_dimensions[get_column_letter(6 + col_offset)].width = 1.68  # Ancho = 1

                # Separador para la siguiente etiqueta a la derecha
                if i < 3:
                    sheet.column_dimensions[get_column_letter(7 + col_offset)].width = 1.55

                # # Ajustar las alturas de las filas
                

                
                
                # Ajustar las alturas de las filas
                sheet.row_dimensions[1 + row_offset].height = 3.8
                sheet.row_dimensions[2 + row_offset].height = 13.5
                sheet.row_dimensions[3 + row_offset].height = 9
                sheet.row_dimensions[4 + row_offset].height = 4.5
                sheet.row_dimensions[5 + row_offset].height = 11.5
                sheet.row_dimensions[6 + row_offset].height = 16
                sheet.row_dimensions[7 + row_offset].height = 29
                sheet.row_dimensions[8 + row_offset].height = 5.3
                # Separador para la siguiente etiqueta abajo
                sheet.row_dimensions[9+ row_offset].height = 8.5

                
                
                
                
                
                # Aplicar bordes manualmente para formar el rectángulo
                # A1: Borde izquierdo y borde arriba
                sheet.cell(row=1 + row_offset, column=1 + col_offset).border = Border(left=thin_border, top=thin_border)

                # A2:A7 Borde izquierdo
                for row in range(2, 8):
                    sheet.cell(row=row + row_offset, column=1 + col_offset).border = Border(left=thin_border)

                # A8 Borde izquierdo y borde abajo
                sheet.cell(row=8 + row_offset, column=1 + col_offset).border = Border(left=thin_border, bottom=thin_border)

                # B1:E1 Borde arriba
                for col in range(2, 6):
                    sheet.cell(row=1 + row_offset, column=col + col_offset).border = Border(top=thin_border)

                # F1: Borde arriba, borde derecho
                sheet.cell(row=1 + row_offset, column=6 + col_offset).border = Border(top=thin_border, right=thin_border)

                # B7:E8 Borde abajo
                for col in range(2, 6):
                    sheet.cell(row=8 + row_offset, column=col + col_offset).border = Border(bottom=thin_border)

                # F8 Borde abajo, borde derecho
                sheet.cell(row=8 + row_offset, column=6 + col_offset).border = Border(bottom=thin_border, right=thin_border)

                # F2:F7 Borde derecho
                for row in range(2, 8):
                    sheet.cell(row=row + row_offset, column=6 + col_offset).border = Border(right=thin_border)

                # Unir celdas B2:E2 y colocar "Nombre de la sustancia"
                sheet.merge_cells(start_row=2 + row_offset, start_column=2 + col_offset, end_row=2 + row_offset, end_column=5 + col_offset)
                merged_cell = sheet.cell(row=2 + row_offset, column=2 + col_offset)
                merged_cell.value = nombre_sustancia.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=10)  # Fuente y tamaño

                # Unir celdas B3:C3 y colocar "CAS"
                sheet.merge_cells(start_row=3 + row_offset, start_column=2 + col_offset, end_row=3 + row_offset, end_column=3 + col_offset)
                merged_cell = sheet.cell(row=3 + row_offset, column=2 + col_offset)
                merged_cell.value = 'CAS'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas D3:E3 y colocar CAS de la sustancia
                sheet.merge_cells(start_row=3 + row_offset, start_column=4 + col_offset, end_row=3 + row_offset, end_column=5 + col_offset)
                merged_cell = sheet.cell(row=3 + row_offset, column=4 + col_offset)
                merged_cell.value = cas.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)  # Formato "Reducir hasta ajustar"
                merged_cell.font = Font(name='Ancizar Sans', size=8)  # Fuente y tamaño

                # Unir celdas B4:B7, girar 90° a la izquierda y colocar palabra de advertencia en mayúscula y negrita
                sheet.merge_cells(start_row=4 + row_offset, start_column=2 + col_offset, end_row=7 + row_offset, end_column=2 + col_offset)
                merged_cell = sheet.cell(row=4 + row_offset, column=2 + col_offset)
                merged_cell.value = advertencia.upper()  # Normalizar a mayúscula
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True, text_rotation=90)  # Girar texto 90°
                merged_cell.font = Font(name='Ancizar Sans', size=9, bold=True)  # Fuente, tamaño y negrita

                # C5: Agregar el pictograma1 si existe
                if sustancia and sustancia.pictogram_clp1:
                    fila_picto=4+row_offset
                    pictogram_path = sustancia.pictogram_clp1.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 40  # Ajuste Altura
                    img.width = 40   # Ajuste Ancho
                    img.anchor = f'{get_column_letter(3 + col_offset)}{fila_picto}'
                    sheet.add_image(img)
                
                # C7: Agregar el pictograma2 si existe
                if sustancia and sustancia.pictogram_clp2:
                    fila_picto=7+row_offset
                    pictogram_path = sustancia.pictogram_clp2.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 40  # Ajuste Altura
                    img.width = 40   # Ajuste Ancho
                    img.anchor = f'{get_column_letter(3 + col_offset)}{fila_picto}'
                    sheet.add_image(img)

                # D6: Agregar el pictograma3 si existe
                if sustancia and sustancia.pictogram_clp3:
                    fila_picto=6+row_offset
                    pictogram_path = sustancia.pictogram_clp3.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 40  # Ajuste Altura
                    img.width = 40   # Ajuste Ancho
                    img.anchor = f'{get_column_letter(4 + col_offset)}{fila_picto}'
                    sheet.add_image(img)

                # E4: Agregar el pictograma4 si existe
                if sustancia and sustancia.pictogram_clp4:
                    fila_picto=4+row_offset
                    pictogram_path = sustancia.pictogram_clp4.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 40  # Ajuste Altura
                    img.width = 40   # Ajuste Ancho
                    img.anchor = f'{get_column_letter(5 + col_offset)}{fila_picto}'
                    sheet.add_image(img)

                # E7: Agregar el pictograma5 si existe
                if sustancia and sustancia.pictogram_clp5:
                    fila_picto=7+row_offset
                    pictogram_path = sustancia.pictogram_clp5.pictogram.path
                    img = openpyxl.drawing.image.Image(pictogram_path)
                    img.height = 40  # Ajuste Altura
                    img.width = 40   # Ajuste Ancho
                    img.anchor = f'{get_column_letter(5 + col_offset)}{fila_picto}'
                    sheet.add_image(img)
        
        # Guardar el archivo en la memoria y preparar la respuesta HTTP
        file_name = get_file_name()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        workbook.save(response)
        return response


class PictogramsList(LoginRequiredMixin, ListView):
    model = Pictograma
    template_name = "Etiquetas/listado_pictogramas.html"
    paginate_by = 10

    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la solicitud del usuario
        per_page = request.GET.get('per_page', request.session.get('per_page', 10))
        self.paginate_by = int(per_page)
        request.session['per_page'] = self.paginate_by
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuarios'] = User.objects.all()
        context['laboratorios'] = Laboratorios.objects.all()
        # Añadir parámetro de ordenamiento a contexto
        context['ordering'] = self.request.GET.get('ordering', '-id')
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        # Obtener el parámetro de ordenamiento desde la solicitud GET
        ordering = self.request.GET.get('ordering', '-id')
        queryset = queryset.order_by(ordering)
        return queryset

# ----------------------- #
# Cargar nuevo pictograma #

class LoadNewPictogram(LoginRequiredMixin, View):
    template_name = 'Etiquetas/crear_pictogramas.html'

    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def get(self, request, *args, **kwargs):
        form = PictogramaForm()
        

        return render(request, self.template_name, {'form': form,})

    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def post(self, request, *args, **kwargs):
        form = PictogramaForm(request.POST, request.FILES)
        try:
            if form.is_valid():
                # Asignar el usuario actual a los campos created_by y last_updated_by
                form.instance.created_by = request.user
                form.instance.last_updated_by = request.user

                
                # Guardar el registro si el formulario es válido
                pictograma = form.save()
                                                
                # Registrar evento
                tipo_evento = 'CARGAR PICTOGRAMA'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = f'Pictograma {pictograma.name} cargado correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                print(form.errors)
                return JsonResponse({'success': False, 'errors': form.errors})
                
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor: {mensaje}')


# --------------------------- #
# Editar pictograma existente #
class EditPictogram(LoginRequiredMixin, View):
    template_name = 'Etiquetas/editar_pictogramas.html'

    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def get(self, request, *args, **kwargs):
        encoded_id = kwargs.get('id')
        decoded_id = base64.b64decode(base64.b64decode(encoded_id)).decode('utf-8')
        pictogram = get_object_or_404(Pictograma, id=decoded_id)
        form = PictogramaForm(instance=pictogram)
        return render(request, self.template_name, {'form': form, 'pictogram': pictogram})

    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def post(self, request, *args, **kwargs):
        encoded_id = kwargs.get('id')
        decoded_id = base64.b64decode(base64.b64decode(encoded_id)).decode('utf-8')
        pictogram = get_object_or_404(Pictograma, id=decoded_id)
        form = PictogramaForm(request.POST, request.FILES, instance=pictogram)
        try:
            if form.is_valid():
                # Asignar el usuario actual a los campos last_updated_by
                form.instance.last_updated_by = request.user

                # Guardar el registro si el formulario es válido
                pictogram = form.save()
                                                
                # Registrar evento
                tipo_evento = 'EDITAR PICTOGRAMA'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

                mensaje = f'Pictograma {pictogram.name} actualizado correctamente.'
                return JsonResponse({'success': True, 'message': mensaje})
            else:
                print(form.errors)
                return JsonResponse({'success': False, 'errors': form.errors})
                
        except Exception as e:
            print(e)
            mensaje = f'Error: {e}'
            return HttpResponseBadRequest(f'Error interno del servidor: {mensaje}')
        
# --------------------- #
# Desactivar Pictograma #
class DisablePictogram(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            pictogram_key = kwargs.get('id')
            item_id_encoded = base64.urlsafe_b64decode(pictogram_key).decode('utf-8')
            pictogram_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia del registro
            pictogram = get_object_or_404(Pictograma, id=pictogram_id)
            # incluir el usuario que realiza el cambio
            pictogram.last_updated_by= request.user
            # Desactivar el registro
            pictogram.is_active = False
            pictogram.save()
            
            # Registrar evento
            tipo_evento = 'DESACTIVAR PICTOGRAMA'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Pictograma {pictogram.name} deshabilitado correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})

# --------------------- #
# Activar Pictograma #
class EnablePictogram(LoginRequiredMixin, View):
    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def post(self, request, *args, **kwargs):
        try:
            # Obtener el ID codificado dos veces desde los parámetros de la solicitud
            pictogram_key = kwargs.get('id')
            item_id_encoded = base64.urlsafe_b64decode(pictogram_key).decode('utf-8')
            pictogram_id = base64.urlsafe_b64decode(item_id_encoded).decode('utf-8')
            
            # Obtener la instancia del registro
            pictogram = get_object_or_404(Pictograma, id=pictogram_id)
            # incluir el usuario que realiza el cambio
            pictogram.last_updated_by= request.user
            # Desactivar el registro
            pictogram.is_active = True
            pictogram.save()
            
            # Registrar evento
            tipo_evento = 'ACTIVAR PICTOGRAMA'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            # Devolver una respuesta JSON de éxito
            return JsonResponse({'success': True, 'message': f'Pictograma {pictogram.name} habilitado correctamente.'})
        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': 'Error interno del servidor'})